# =========================================================
# CODEXUP PREMIUM TERMINAL — FINANCE SUITE (DOUBLE-ENTRY)
# VERSION TIMESTAMP: 2026-06-06 04:00 PM (ACCOUNTING UPGRADE)
# =========================================================

import io
import json
import os
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import yfinance as yf
from dateutil.relativedelta import relativedelta
import requests

# Import openpyxl engines for programmatic Excel generation
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

# -----------------------------
# LOCAL PERSISTENCE — AUTO-SAVE / AUTO-LOAD
# Data is stored in finance_data.json next to this script
# -----------------------------
DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "finance_data.json")

DEFAULT_COA = {
    # Assets
    "Cash": "Asset",
    "Punjab National Bank": "Asset",
    "Jio Payment Bank": "Asset",
    "Stock Market Asset": "Asset",
    "Gold Asset": "Asset",
    "BlinkX Account": "Asset",
    "INDmoney Account": "Asset",
    "Mutual Fund": "Asset",
    "KSFE Sugama Account 004000010006313": "Asset",
    # Liabilities
    "Bajaj Loan": "Liability",
    "PNB Gold Loan": "Liability",
    "KSFE Gold Loan": "Liability",
    "Chitty Liability": "Liability",
    "Credit Card": "Liability",
    "KSFE GOLD LOAN 00400120052099": "Liability",
    "KSFE GOLD LOAN 00400120052097": "Liability",
    "KSFE GOLD LOAN 00400120052100": "Liability",
    "Accrued Interest(KSFE Gold Loan)": "Liability",
    "Accrued Interest(PNB Gold Loan)": "Liability",
    # Equity
    "Retained Earnings": "Equity",
    "Opening Balance Equity": "Equity",
    # Revenue
    "Salary": "Revenue",
    "Stock Market Gains": "Revenue",
    "Bank Interest Received": "Revenue",
    "Trading Income": "Revenue",
    # Expense
    "Gold Loan PNB Interest": "Expense",
    "Chitty": "Expense",
    "KSFE Interest Tiers": "Expense",
    "Innamma": "Expense",
    "Other Expenses": "Expense",
    "Stock Market Loss": "Expense",
    "Groceries": "Expense",
    "Refershment": "Expense",
    "Travelling Expense(Petrol)": "Expense",
    "Travelling Expense(Radhu)": "Expense",
    "Shopping": "Expense",
    "Mobile Expenses": "Expense",
    "Bank Charges": "Expense",
    "Bank Interest": "Expense",
    "Trading Loss": "Expense",
    "Subscription": "Expense",
    "KSFE Gold Loan Interest": "Expense"
}

def on_quick_narration_change(suggestion_key):
    val = st.session_state[suggestion_key]
    if val != "-- Select to reuse narration --":
        st.session_state.quick_narration_val = val
        # Sync with Streamlit widget state key directly
        v = st.session_state.get("qe_form_version", 0)
        st.session_state[f"qe_narration_input_{v}"] = val
        st.session_state[suggestion_key] = "-- Select to reuse narration --"

def on_adv_narration_change(suggestion_key):
    val = st.session_state[suggestion_key]
    if val != "-- Select to reuse narration --":
        st.session_state.adv_narration_val = val
        # Sync with Streamlit widget state key directly
        v = st.session_state.get("adv_form_version", 0)
        st.session_state[f"adv_narration_input_{v}"] = val
        st.session_state[suggestion_key] = "-- Select to reuse narration --"

def on_edit_narration_change(state_narr_key, suggestion_key):
    val = st.session_state[suggestion_key]
    if val != "-- Select to reuse narration --":
        st.session_state[state_narr_key] = val
        # Sync with Streamlit widget state key directly
        target_jv_id = state_narr_key[19:]
        widget_key = f"inline_edit_narr_{target_jv_id}"
        st.session_state[widget_key] = val
        st.session_state[suggestion_key] = "-- Select to reuse narration --"

def get_account_type(acc_name):
    """Retrieve the type of an account, defaulting to Asset if unknown."""
    if not acc_name:
        return "Asset"
    acc = st.session_state.accounts.get(acc_name)
    if isinstance(acc, dict):
        return acc.get("type", "Asset")
    return acc or "Asset"

def get_account_parent(acc_name):
    """Retrieve the parent account of an account, or None if it has no parent."""
    if not acc_name:
        return None
    acc = st.session_state.accounts.get(acc_name)
    if isinstance(acc, dict):
        return acc.get("parent")
    return None

def on_edit_target_change():
    target = st.session_state.sel_edit_target_acc
    if target:
        st.session_state.txt_edit_acc_name = target
        target_type = get_account_type(target)
        st.session_state.sel_edit_acc_type = target_type
        st.session_state.sel_edit_acc_parent = get_account_parent(target)

def on_edit_type_change():
    st.session_state.sel_edit_acc_parent = None

def on_add_type_change():
    st.session_state.sel_add_acc_parent = None

def to_date_obj(d):
    if isinstance(d, str):
        try:
            return datetime.strptime(d, "%Y-%m-%d").date()
        except ValueError:
            return None
    if isinstance(d, datetime):
        return d.date()
    return d

def get_account_start_date(account_name):
    first_date = None
    for jv in st.session_state.journal_entries:
        for line in jv["lines"]:
            if line["account"] == account_name:
                jv_date = to_date_obj(jv["date"])
                if jv_date:
                    if first_date is None or jv_date < first_date:
                        first_date = jv_date
    return first_date or datetime.now().date()

def get_monthly_payment(account_name):
    payments_by_month = {}
    for jv in st.session_state.journal_entries:
        jv_date = to_date_obj(jv["date"])
        if not jv_date:
            continue
        month_key = jv_date.strftime("%Y-%m")
        for line in jv["lines"]:
            if line["account"] == account_name and line["debit"] > 0:
                payments_by_month[month_key] = payments_by_month.get(month_key, 0.0) + line["debit"]
    
    if not payments_by_month:
        return 0.0
    sorted_months = sorted(payments_by_month.keys())
    return payments_by_month[sorted_months[-1]]

# -----------------------------
# ACCOUNT AND LEDGER HELPER ROUTINES
# -----------------------------
@st.cache_data(ttl=3600)
def fetch_gold_price_inr():
    if "cached_gold_rate" in st.session_state and st.session_state.cached_gold_rate is not None:
        return st.session_state.cached_gold_rate

    price = None
    # 1. Try scraping Kerala gold price from GoodReturns (most realistic for local rate)
    try:
        import urllib.request
        import re
        url = "https://www.goodreturns.in/gold-rates/kerala.html"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=5) as response:
            html = response.read().decode('utf-8', errors='ignore')
        
        # We prefer 22K gold rate as it represents standard jewelry/loan valuation rate
        match_22 = re.search(r'id="22K-price"[^>]*>&#x20b9;([\d,]+)</span>', html)
        if match_22:
            price_str = match_22.group(1).replace(",", "")
            price = float(price_str)
        else:
            # Fallback to 24K gold rate if 22K is missing
            match_24 = re.search(r'id="24K-price"[^>]*>&#x20b9;([\d,]+)</span>', html)
            if match_24:
                price_str = match_24.group(1).replace(",", "")
                price = float(price_str)
    except Exception:
        pass

    if price is None:
        # 2. If scraping fails, try fetching from yfinance (global spot price conversion)
        try:
            gold = yf.Ticker("GC=F")
            gold_price_usd = gold.history(period="1d")["Close"].iloc[-1]
            inr = yf.Ticker("INR=X")
            inr_rate = inr.history(period="1d")["Close"].iloc[-1]
            price_per_gram_inr = (gold_price_usd / 31.1034768) * inr_rate
            price = price_per_gram_inr
        except Exception:
            pass

    if price is None:
        # 3. Ultimate static fallback (approximate recent market rate)
        price = 14000.0

    st.session_state.cached_gold_rate = price
    return price

def get_all_journal_entries():
    jvs = list(st.session_state.journal_entries)
    
    # Calculate dynamic gold valuation
    gold_rate = fetch_gold_price_inr()
    gold_qty = st.session_state.get("gold_qty", 177.0)
    gold_depreciation = 0.23
    val = gold_qty * gold_rate * (1 - gold_depreciation)
    
    # Only add valuation if we have gold qty
    if val > 0:
        jvs.append({
            "jv_id": "JV-VAL-GOLD",
            "date": datetime(2026, 6, 6).date(),
            "narration": f"Dynamic Gold Asset Valuation ({gold_qty:,.2f}g @ ₹{gold_rate:,.2f}/g less {gold_depreciation * 100:.0f}%)",
            "lines": [
                {"account": "Gold Asset", "debit": val, "credit": 0.0},
                {"account": "Opening Balance Equity", "debit": 0.0, "credit": val}
            ]
        })
    return jvs

def get_account_balance(account_name):
    """Calculate the net balance of an account based on its normal type rules."""
    account_type = get_account_type(account_name)
    total_dr = 0.0
    total_cr = 0.0
    for jv in get_all_journal_entries():
        for line in jv["lines"]:
            if line["account"] == account_name:
                total_dr += line["debit"]
                total_cr += line["credit"]
                
    if account_type in ["Asset", "Expense"]:
        return total_dr - total_cr
    else:
        return total_cr - total_dr

def get_rolled_up_balances(level_view="Detailed"):
    """Calculate balances for accounts, rolling up child accounts if level_view is 'Main Ledgers Only'."""
    raw_balances = {}
    for name in st.session_state.accounts:
        raw_balances[name] = get_account_balance(name)
        
    if level_view == "Main Ledgers Only":
        rolled = {}
        for name in st.session_state.accounts:
            parent = get_account_parent(name)
            typ = get_account_type(name)
            bal = raw_balances[name]
            
            if parent and parent in st.session_state.accounts:
                if parent not in rolled:
                    rolled[parent] = {
                        "balance": 0.0,
                        "type": get_account_type(parent)
                    }
                rolled[parent]["balance"] += bal
            else:
                if name not in rolled:
                    rolled[name] = {
                        "balance": 0.0,
                        "type": typ
                    }
                rolled[name]["balance"] += bal
        return rolled
    else:
        detailed = {}
        for name in st.session_state.accounts:
            detailed[name] = {
                "balance": raw_balances[name],
                "type": get_account_type(name)
            }
        return detailed

def get_ledger_for_account(account_name):
    """Retrieve chronologically ordered general ledger transactions with running balance."""
    account_type = get_account_type(account_name)
    rows = []
    
    for jv in get_all_journal_entries():
        # Check if the JV involves the chosen account
        matching_lines = [l for l in jv["lines"] if l["account"] == account_name]
        if not matching_lines:
            continue
            
        other_accounts = list(set([l["account"] for l in jv["lines"] if l["account"] != account_name]))
        opposing = ", ".join(other_accounts) if other_accounts else "Self-balancing"
        
        for l in matching_lines:
            rows.append({
                "Date": jv["date"],
                "JV ID": jv["jv_id"],
                "Narration": jv["narration"],
                "Opposing Account": opposing,
                "Debit": l["debit"],
                "Credit": l["credit"]
            })
            
    df_ledger = pd.DataFrame(rows)
    if df_ledger.empty:
        return pd.DataFrame(columns=["Date", "JV ID", "Narration", "Opposing Account", "Debit", "Credit", "Running Balance"])
        
    df_ledger = df_ledger.sort_values(by=["Date", "JV ID"]).reset_index(drop=True)
    
    bal = 0.0
    balances = []
    for _, row in df_ledger.iterrows():
        dr = row["Debit"]
        cr = row["Credit"]
        if account_type in ["Asset", "Expense"]:
            bal += dr - cr
        else:
            bal += cr - dr
        balances.append(bal)
        
    df_ledger["Running Balance"] = balances
    return df_ledger

def migrate_accounts(accounts_dict):
    """Migrate flat accounts dictionary schema to the hierarchical schema."""
    migrated = {}
    for name, val in accounts_dict.items():
        if isinstance(val, str):
            migrated[name] = {"type": val, "parent": None}
        elif isinstance(val, dict):
            migrated[name] = {
                "type": val.get("type", "Asset"),
                "parent": val.get("parent")
            }
        else:
            migrated[name] = {"type": "Asset", "parent": None}
    return migrated

def format_account_label(acc_name):
    """Format account name for selectbox dropdowns to show visual hierarchy."""
    if acc_name is None:
        return ""
    parent = get_account_parent(acc_name)
    if parent:
        return f"{parent} ➔ {acc_name}"
    return acc_name

def load_from_cloud(url, secret):
    try:
        base_url = url.rstrip("/")
        if not base_url.endswith(".json"):
            fetch_url = f"{base_url}/ledger.json"
        else:
            fetch_url = base_url
            
        params = {}
        if secret:
            params["auth"] = secret
            
        response = requests.get(fetch_url, params=params, timeout=5)
        if response.status_code == 200:
            return response.json()
    except Exception as e:
        st.warning(f"⚠️ Could not sync with cloud database: {e}")
    return None

def save_to_cloud(payload, url, secret):
    try:
        base_url = url.rstrip("/")
        if not base_url.endswith(".json"):
            write_url = f"{base_url}/ledger.json"
        else:
            write_url = base_url
            
        params = {}
        if secret:
            params["auth"] = secret
            
        response = requests.put(write_url, json=payload, params=params, timeout=5)
        return response.status_code == 200
    except Exception as e:
        pass
    return False

def save_data():
    """Persist all double-entry app data to the local JSON file."""
    try:
        entries_to_save = []
        for entry in st.session_state.journal_entries:
            e = entry.copy()
            if hasattr(e["date"], "strftime"):
                e["date"] = e["date"].strftime("%Y-%m-%d")
            e["lines"] = [
                {
                    "account": line["account"],
                    "debit": float(line["debit"]),
                    "credit": float(line["credit"])
                }
                for line in e["lines"]
            ]
            entries_to_save.append(e)
            
        payload = {
            "owner_name": st.session_state.get("owner_name", "Abhilash"),
            "accounts": st.session_state.accounts,
            "journal_entries": entries_to_save,
            "gold_qty": st.session_state.get("gold_qty", 177.0),
            "password": st.session_state.get("app_password", "finance@2026"),
            "security_question": st.session_state.get("security_question", "What is the owner name of this finance ledger?"),
            "security_answer": st.session_state.get("security_answer", "Abhilash"),
            "cloud_sync_enabled": st.session_state.get("cloud_sync_enabled", False),
            "cloud_url": st.session_state.get("cloud_url", ""),
            "cloud_secret": st.session_state.get("cloud_secret", ""),
            "important_events": st.session_state.get("important_events", [])
        }
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
            
        if st.session_state.get("cloud_sync_enabled", False) and st.session_state.get("cloud_url", ""):
            save_to_cloud(payload, st.session_state.cloud_url, st.session_state.cloud_secret)
    except Exception as e:
        st.error(f"❌ Error saving data: {e}")

def load_data():
    """Restore double-entry app data and run legacy migration if required."""
    if not os.path.exists(DATA_FILE):
        st.session_state.accounts = migrate_accounts(DEFAULT_COA)
        st.session_state.journal_entries = []
        st.session_state.owner_name = "Abhilash"
        st.session_state.gold_qty = 177.0
        st.session_state.cloud_sync_enabled = False
        st.session_state.cloud_url = ""
        st.session_state.cloud_secret = ""
        st.session_state.important_events = []
        return False
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            d = json.load(f)

        # Load cloud settings from local file
        st.session_state.cloud_sync_enabled = d.get("cloud_sync_enabled", False)
        st.session_state.cloud_url = d.get("cloud_url", "")
        st.session_state.cloud_secret = d.get("cloud_secret", "")

        # Try syncing from Cloud if enabled
        if st.session_state.cloud_sync_enabled and st.session_state.cloud_url:
            cloud_d = load_from_cloud(st.session_state.cloud_url, st.session_state.cloud_secret)
            if cloud_d:
                d = cloud_d

        # ─ 1. Detect Legacy Single-Entry Format & Migrate ───────────────────
        if "transaction_data" in d and "journal_entries" not in d:
            accounts = DEFAULT_COA.copy()
            
            # Extract legacy properties
            legacy_cats = d.get("valid_categories", [])
            income_cats = d.get("income_categories", [])
            
            # Map legacy categories to account types
            for cat in legacy_cats:
                if cat not in accounts:
                    if cat in income_cats:
                        accounts[cat] = "Revenue"
                    else:
                        accounts[cat] = "Expense"
            
            journal_entries = []
            legacy_txs = d.get("transaction_data", [])
            for idx, tx in enumerate(legacy_txs, start=1):
                jv_id = f"JV-{idx:05d}"
                date_str = tx.get("Date", datetime.now().strftime("%Y-%m-%d"))
                desc = tx.get("Description", "Migrated Entry")
                category = tx.get("Category", "Other Expenses")
                amount = float(tx.get("Amount", 0.0))
                
                if category not in accounts:
                    accounts[category] = "Expense"
                
                # Formulate debit and credit legs
                lines = []
                if category in income_cats or accounts.get(category) == "Revenue":
                    lines.append({"account": "Punjab National Bank", "debit": amount, "credit": 0.0})
                    lines.append({"account": category, "debit": 0.0, "credit": amount})
                else:
                    lines.append({"account": category, "debit": amount, "credit": 0.0})
                    lines.append({"account": "Punjab National Bank", "debit": 0.0, "credit": amount})
                
                try:
                    dt = datetime.strptime(date_str, "%Y-%m-%d").date()
                except Exception:
                    dt = datetime.now().date()
                    
                journal_entries.append({
                    "jv_id": jv_id,
                    "date": dt,
                    "narration": desc,
                    "lines": lines
                })
                
            st.session_state.accounts = migrate_accounts(accounts)
            st.session_state.journal_entries = journal_entries
            st.session_state.owner_name = "Abhilash"
            st.session_state.gold_qty = 177.0
            st.session_state.app_password = "finance@2026"
            st.session_state.security_question = "What is the owner name of this finance ledger?"
            st.session_state.security_answer = "Abhilash"
            st.session_state.important_events = []
            
            save_data() # Persist migrated structure immediately
            return True

        # ─ 2. Standard Double-Entry Load ────────────────────────────────────
        st.session_state.accounts = migrate_accounts(d.get("accounts", DEFAULT_COA))
        st.session_state.owner_name = d.get("owner_name", "Abhilash")
        st.session_state.gold_qty = d.get("gold_qty", 177.0)
        st.session_state.app_password = d.get("password", "finance@2026")
        st.session_state.security_question = d.get("security_question", "What is the owner name of this finance ledger?")
        st.session_state.security_answer = d.get("security_answer", "Abhilash")
        st.session_state.cloud_sync_enabled = d.get("cloud_sync_enabled", False)
        st.session_state.cloud_url = d.get("cloud_url", "")
        st.session_state.cloud_secret = d.get("cloud_secret", "")
        st.session_state.important_events = d.get("important_events", [])
        
        # Check and rename Bank Account -> Punjab National Bank, add Jio Payment Bank
        if "Bank Account" in st.session_state.accounts:
            st.session_state.accounts["Punjab National Bank"] = {"type": "Asset", "parent": None}
            st.session_state.accounts.pop("Bank Account", None)
            
        if "Jio Payment Bank" not in st.session_state.accounts:
            st.session_state.accounts["Jio Payment Bank"] = {"type": "Asset", "parent": None}
            
        if "Punjab National Bank" not in st.session_state.accounts:
            st.session_state.accounts["Punjab National Bank"] = {"type": "Asset", "parent": None}
            
        if "Cash" not in st.session_state.accounts:
            st.session_state.accounts["Cash"] = {"type": "Asset", "parent": None}
            
        if "BlinkX Account" not in st.session_state.accounts:
            st.session_state.accounts["BlinkX Account"] = {"type": "Asset", "parent": None}
            
        if "INDmoney Account" not in st.session_state.accounts:
            st.session_state.accounts["INDmoney Account"] = {"type": "Asset", "parent": None}
            
        jes = []
        for entry in d.get("journal_entries", []):
            je = entry.copy()
            if "date" in je:
                je["date"] = datetime.strptime(je["date"], "%Y-%m-%d").date()
            # Upgrade legacy Bank Account to Punjab National Bank in line records
            for line in je.get("lines", []):
                if line["account"] == "Bank Account":
                    line["account"] = "Punjab National Bank"
            jes.append(je)
        st.session_state.journal_entries = jes
        
        # Save structural upgrades
        save_data()
        return True
    except Exception:
        # Fallback initialization
        st.session_state.accounts = migrate_accounts(DEFAULT_COA)
        st.session_state.journal_entries = []
        st.session_state.owner_name = "Abhilash"
        st.session_state.gold_qty = 177.0
        st.session_state.app_password = "finance@2026"
        st.session_state.security_question = "What is the owner name of this finance ledger?"
        st.session_state.security_answer = "Abhilash"
        st.session_state.cloud_sync_enabled = False
        st.session_state.cloud_url = ""
        st.session_state.cloud_secret = ""
        st.session_state.important_events = []
        return False

# -----------------------------
# PAGE CONFIG & THEME INITIALIZATION
# -----------------------------
st.set_page_config(
    page_title="Personal Finance Manager",
    layout="wide",
    initial_sidebar_state="expanded"
)

# -----------------------------
# INITIALIZE STATE DATA (Loaded early to supply owner name)
# -----------------------------
if "_ledger_loaded" not in st.session_state:
    st.session_state["_ledger_loaded"] = load_data()

# Initialize login state
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    # Render centered login block using columns
    _, col_login, _ = st.columns([1, 1.2, 1])
    with col_login:
        st.markdown("<div style='height:80px'></div>", unsafe_allow_html=True)
        
        # Decide if showing forgot password screen or main login screen
        if "forgot_password_mode" not in st.session_state:
            st.session_state.forgot_password_mode = False
            
        if st.session_state.forgot_password_mode:
            with st.container(border=True):
                st.markdown(f"""
                <div style='text-align: center; margin-bottom: 20px;'>
                    <h2 style='margin: 0; font-family: "Outfit", sans-serif; font-size: 1.6rem; font-weight: 800; color: #0F172A;'>🔑 Reset Password</h2>
                    <p style='margin: 4px 0 0 0; font-size: 0.85rem; color: #64748B;'>Answer your security recovery question.</p>
                </div>
                """, unsafe_allow_html=True)
                
                st.info(f"**Question**: {st.session_state.get('security_question', 'What is the owner name of this finance ledger?')}")
                answer_input = st.text_input("Answer", value="", placeholder="Enter answer...", key="reset_answer_input")
                new_pw_input = st.text_input("New Password", type="password", value="", placeholder="Enter new password...", key="reset_new_pw_input")
                confirm_pw_input = st.text_input("Confirm New Password", type="password", value="", placeholder="Confirm new password...", key="reset_confirm_pw_input")
                
                st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
                
                col_reset1, col_reset2 = st.columns(2)
                with col_reset1:
                    if st.button("Reset Password", use_container_width=True, type="primary", key="btn_reset_submit"):
                        correct_ans = st.session_state.get("security_answer", "Abhilash").strip().lower()
                        if answer_input.strip().lower() == correct_ans:
                            if not new_pw_input:
                                st.error("⚠️ Password cannot be empty.")
                            elif new_pw_input != confirm_pw_input:
                                st.error("⚠️ Passwords do not match.")
                            else:
                                st.session_state.app_password = new_pw_input
                                save_data()
                                st.session_state.forgot_password_mode = False
                                st.session_state.login_success_msg = "✅ Password reset successful! Please login."
                                st.rerun()
                        else:
                            st.error("⚠️ Incorrect answer to security question.")
                with col_reset2:
                    if st.button("Back to Login", use_container_width=True, key="btn_reset_back"):
                        st.session_state.forgot_password_mode = False
                        st.rerun()
        else:
            with st.container(border=True):
                st.markdown("""
                <div style='text-align: center; margin-bottom: 20px;'>
                    <h2 style='margin: 0; font-family: "Outfit", sans-serif; font-size: 1.6rem; font-weight: 800; color: #0F172A;'>🔐 FinPlus Security</h2>
                    <p style='margin: 4px 0 0 0; font-size: 0.85rem; color: #64748B;'>Please enter your credentials to unlock the ledger.</p>
                </div>
                """, unsafe_allow_html=True)
                
                if "login_success_msg" in st.session_state and st.session_state.login_success_msg:
                    st.success(st.session_state.login_success_msg)
                    del st.session_state.login_success_msg
                
                login_user = st.text_input("Username", value="", placeholder="Username (case-insensitive)", key="login_username_input")
                login_pw = st.text_input("Password", value="", type="password", placeholder="Password", key="login_password_input")
                
                st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
                
                col_btn1, col_btn2 = st.columns([2, 1.2])
                with col_btn1:
                    if st.button("Unlock Ledger", use_container_width=True, type="primary", key="btn_login_submit"):
                        # Compare with app_password from session state (loaded from JSON)
                        correct_user = st.session_state.get("owner_name", "Abhilash").strip().lower()
                        correct_pw = st.session_state.get("app_password", "finance@2026")
                        if login_user.strip().lower() == correct_user and login_pw == correct_pw:
                            st.session_state.logged_in = True
                            st.success("🔑 Access granted! Loading...")
                            st.rerun()
                        else:
                            st.error("⚠️ Invalid username or password.")
                with col_btn2:
                    if st.button("Forgot?", use_container_width=True, key="btn_forgot_trigger"):
                        st.session_state.forgot_password_mode = True
                        st.rerun()
        st.stop()

if "qe_form_version" not in st.session_state:
    st.session_state.qe_form_version = 0
if "adv_form_version" not in st.session_state:
    st.session_state.adv_form_version = 0

# -----------------------------
# BRAND HEADER BANNER
# -----------------------------
logo_path = r"C:\Users\AbhilashBabu\Finance\finplus_brand_image.png"

# Calculate Current Net Worth dynamically
asset_accounts = [name for name in st.session_state.accounts if get_account_type(name) == "Asset"]
other_assets_sum = sum(get_account_balance(name) for name in asset_accounts if name != "Gold Asset")
gold_rate = fetch_gold_price_inr()
gold_qty = st.session_state.get("gold_qty", 177.0)
gold_depreciation = 0.23
live_gold_bal = get_account_balance("Gold Asset")
current_gold_value = live_gold_bal if live_gold_bal > 0 else (gold_qty * gold_rate * (1 - gold_depreciation))
total_assets_now = other_assets_sum + current_gold_value

liab_accounts = [name for name in st.session_state.accounts if get_account_type(name) == "Liability"]
total_liabilities_now = sum(get_account_balance(name) for name in liab_accounts)

net_worth_now = total_assets_now - total_liabilities_now

header_col1, header_col2, header_col3 = st.columns([1.2, 6.8, 2.0])
with header_col1:
    if os.path.exists(logo_path):
        st.image(logo_path, use_container_width=True)
    else:
        st.markdown("<div style='font-size:3rem;'>📊</div>", unsafe_allow_html=True)

with header_col2:
    owner = st.session_state.get("owner_name", "Abhilash")
    st.markdown(f"""
    <div style="padding-top: 10px;">
        <h1 style="margin: 0; font-family: 'Outfit', sans-serif; font-size: 2.4rem; font-weight: 800; color: #0F172A; letter-spacing: -0.03em;">
            Personal Finance Manager
        </h1>
        <div style="font-family: 'Plus Jakarta Sans', sans-serif; font-size: 1.05rem; color: #475569; font-weight: 600; margin-top: 4px; display: flex; flex-direction: column; gap: 4px;">
            <div>Owner Name: <span style="color: #059669; font-weight: 700;">{owner}</span></div>
            <div style="margin-top: 2px;">Current Net Worth: <span style="color: #10B981; font-weight: 800;">₹{net_worth_now:,.2f}</span></div>
        </div>
    </div>
    """, unsafe_allow_html=True)

with header_col3:
    st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)
    if st.button("🔒 Logout", use_container_width=True, key="btn_logout"):
        st.session_state.logged_in = False
        st.rerun()

st.markdown("<div style='height:15px'></div>", unsafe_allow_html=True)

# Late-injected CSS for premium look and feel
st.markdown("""
<style>
/* ─── FONTS ─────────────────────────────────────────────────────────── */
@import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&family=Plus+Jakarta+Sans:wght@300;400;500;600;700&display=swap');

/* ─── BASE ──────────────────────────────────────────────────────────── */
html, body, [class*="css"] {
    font-family: 'Plus Jakarta Sans', 'Outfit', sans-serif !important;
}
.stApp {
    background-color: #F8FAFC !important;
    color: #0F172A !important;
}
.block-container {
    padding: 1.5rem 2rem 3rem !important;
    max-width: 1400px !important;
}

/* ─── SIDEBAR ───────────────────────────────────────────────────────── */
section[data-testid="stSidebar"] {
    background-color: #FFFFFF !important;
    border-right: 1px solid #E2E8F0 !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important;
}
section[data-testid="stSidebar"] * {
    color: #334155 !important;
}
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {
    color: #059669 !important;
    font-size: 0.78rem !important;
    font-weight: 700 !important;
    letter-spacing: 0.12em !important;
    text-transform: uppercase !important;
    margin: 1.4rem 0 0.5rem !important;
    padding-bottom: 0.35rem !important;
    border-bottom: 1px solid #E2E8F0 !important;
}

/* ─── METRIC CARDS ──────────────────────────────────────────────────── */
[data-testid="stMetric"] {
    background: #FFFFFF !important;
    border: 1px solid #E2E8F0 !important;
    border-top: 3px solid #10B981 !important;
    border-bottom: 4px solid #CBD5E1 !important;
    border-radius: 12px !important;
    padding: 16px 20px 14px !important;
    box-shadow: 0 6px 12px -3px rgba(0,0,0,0.03) !important;
    transition: all 0.3s cubic-bezier(0.4,0,0.2,1) !important;
}
[data-testid="stMetric"]:hover {
    border-color: #A7F3D0 !important;
    border-bottom-color: #10B981 !important;
    box-shadow: 0 14px 22px -6px rgba(16,185,129,0.15) !important;
    transform: translateY(-4px) !important;
}
div[data-testid="stMetricLabel"] {
    font-size: 0.78rem !important;
    font-weight: 700 !important;
    color: #64748B !important;
    text-transform: uppercase;
    letter-spacing: 0.09em;
}
div[data-testid="stMetricValue"] {
    font-family: 'Outfit', sans-serif !important;
    font-size: 1.8rem !important;
    font-weight: 700 !important;
    color: #0F172A !important;
}

/* ─── BUTTONS ───────────────────────────────────────────────────────── */
div.stButton > button,
div.stDownloadButton > button {
    background: linear-gradient(to bottom, #FFFFFF, #F8FAFC) !important;
    color: #0F172A !important;
    border-left: 1px solid #E2E8F0 !important;
    border-right: 1px solid #E2E8F0 !important;
    border-top: 1px solid #E2E8F0 !important;
    border-bottom: 4.5px solid #CBD5E1 !important;
    border-radius: 10px !important;
    font-family: 'Outfit', sans-serif !important;
    font-weight: 700 !important;
    font-size: 0.88rem !important;
    padding: 10px 24px !important;
    box-shadow: 0 4px 6px rgba(0,0,0,0.03) !important;
    transition: all 0.15s ease !important;
}
div.stButton > button:hover,
div.stDownloadButton > button:hover {
    background: linear-gradient(to bottom, #10B981, #059669) !important;
    color: #FFFFFF !important;
    border-color: #059669 !important;
    border-bottom-color: #047857 !important;
    box-shadow: 0 6px 16px rgba(16,185,129,0.35) !important;
    transform: translateY(-2px) !important;
}
div.stButton > button:active,
div.stDownloadButton > button:active {
    border-bottom-width: 1px !important;
    transform: translateY(3.5px) !important;
    box-shadow: 0 1px 3px rgba(16,185,129,0.2) !important;
}
div.stButton > button:disabled,
div.stDownloadButton > button:disabled {
    background: #E2E8F0 !important;
    color: #94A3B8 !important;
    border-color: #CBD5E1 !important;
    border-bottom: 1px solid #CBD5E1 !important;
    box-shadow: none !important;
    cursor: not-allowed !important;
    transform: none !important;
}

/* ─── TABS ──────────────────────────────────────────────────────────── */
div[data-baseweb="tab-list"] {
    background: #F1F5F9 !important;
    border-radius: 12px !important;
    padding: 4px !important;
    border: 1px solid #E2E8F0 !important;
    gap: 0 !important;
}
button[data-baseweb="tab"] {
    font-family: 'Outfit', sans-serif !important;
    font-size: 0.92rem !important;
    font-weight: 500 !important;
    padding: 0.6rem 1.25rem !important;
    color: #475569 !important;
    background: transparent !important;
    border: none !important;
    border-radius: 8px !important;
    transition: all 0.2s ease !important;
}
button[data-baseweb="tab"]:hover { color: #0F172A !important; }
button[data-baseweb="tab"][aria-selected="true"] {
    background: #FFFFFF !important;
    color: #0F172A !important;
    font-weight: 600 !important;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05) !important;
}

/* ─── INPUTS ────────────────────────────────────────────────────────── */
.stTextInput input,
.stNumberInput input,
.stDateInput input,
textarea {
    background: #FFFFFF !important;
    color: #0F172A !important;
    border: 1px solid #CBD5E1 !important;
    border-radius: 10px !important;
}
div[data-baseweb="base-input"],
.stNumberInput div[data-baseweb="input"] {
    background: #FFFFFF !important;
    border: 1px solid #CBD5E1 !important;
    border-radius: 10px !important;
}
div[data-baseweb="select"] > div,
div[data-baseweb="select"] > div > div {
    background: #FFFFFF !important;
    color: #0F172A !important;
    border: 1px solid #CBD5E1 !important;
    border-radius: 10px !important;
}
div[data-baseweb="select"] * { color: #0F172A !important; }
div[data-baseweb="select"] svg { fill: #64748B !important; }
input::placeholder { color: #94A3B8 !important; opacity: 1 !important; }

/* ─── DATAFRAME / DATA EDITOR ───────────────────────────────────────── */
div[data-testid="stDataFrame"],
div[data-testid="stTable"] {
    border: 1px solid #E2E8F0 !important;
    border-radius: 8px !important;
    overflow: hidden !important;
}
[data-testid="stDataEditor"] {
    border-radius: 12px !important;
    border: 1px solid #E2E8F0 !important;
    overflow: hidden !important;
}

/* ─── FILE UPLOADER ─────────────────────────────────────────────────── */
section[data-testid="stFileUploader"],
section[data-testid="stFileUploader"] > div {
    background: #FFFFFF !important;
    border: 1px solid #E2E8F0 !important;
    border-radius: 12px !important;
    padding: 4px !important;
}
section[data-testid="stFileUploader"] span,
section[data-testid="stFileUploader"] p { color: #475569 !important; }
section[data-testid="stFileUploaderDropzone"] {
    background: #FFFFFF !important;
    border: 2px dashed #CBD5E1 !important;
    border-radius: 12px !important;
}
section[data-testid="stFileUploader"] svg { fill: #64748B !important; }

/* ─── HERO PANEL ────────────────────────────────────────────────────── */
.hero-panel {
    background: linear-gradient(135deg, #0F172A 0%, #1E293B 100%) !important;
    border: 1px solid #334155 !important;
    border-left: 5px solid #10B981 !important;
    border-radius: 16px !important;
    padding: 1.5rem 2rem !important;
    margin-bottom: 1.5rem !important;
    box-shadow: 0 10px 25px -5px rgba(15, 23, 42, 0.15), 0 8px 10px -6px rgba(15, 23, 42, 0.15) !important;
    position: relative;
    overflow: hidden;
}
.hero-title {
    color: #FFFFFF !important;
    font-family: 'Outfit', sans-serif !important;
    font-weight: 800 !important;
    font-size: 1.8rem !important;
    margin: 0 !important;
    background: linear-gradient(135deg, #FFFFFF 40%, #A7F3D0 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    letter-spacing: -0.02em;
}
.hero-icon {
    font-size: 1.8rem;
    background: rgba(16, 185, 129, 0.12);
    padding: 8px 12px;
    border-radius: 12px;
    display: flex;
    align-items: center;
    justify-content: center;
    border: 1px solid rgba(16, 185, 129, 0.25);
    box-shadow: 0 0 15px rgba(16, 185, 129, 0.15);
    animation: float 3s ease-in-out infinite;
}
@keyframes float {
    0% { transform: translateY(0px); }
    50% { transform: translateY(-4px); }
    100% { transform: translateY(0px); }
}
.hero-glow {
    position: absolute;
    top: -50%;
    right: -20%;
    width: 300px;
    height: 300px;
    background: radial-gradient(circle, rgba(16, 185, 129, 0.12) 0%, rgba(0,0,0,0) 70%);
    z-index: 1;
    pointer-events: none;
}
.hero-badge {
    display: inline-block;
    background: rgba(16, 185, 129, 0.12);
    color: #34D399 !important;
    font-size: 0.72rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    padding: 3px 8px;
    border-radius: 20px;
    margin-top: 4px;
    border: 1px solid rgba(16, 185, 129, 0.25);
}

/* ─── GENERAL TEXT ──────────────────────────────────────────────────── */
h1, h2, h3 {
    color: #0F172A !important;
    font-family: 'Outfit', sans-serif !important;
    font-weight: 700 !important;
}
label, p { color: #475569 !important; }
strong { color: #0F172A !important; }
hr { border-color: #E2E8F0 !important; }

/* ─── PRIMARY / FORM SUBMIT BUTTON ──────────────────────────────────── */
button[kind="primary"],
button[kind="primaryFormSubmit"],
button[data-testid="baseButton-primary"],
div[data-testid="stFormSubmitButton"] > button {
    background: linear-gradient(to bottom, #059669, #047857) !important;
    color: #FFFFFF !important;
    border-left: 1px solid #065F46 !important;
    border-right: 1px solid #065F46 !important;
    border-top: 1px solid #065F46 !important;
    border-bottom: 4.5px solid #064E3B !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    font-size: 0.88rem !important;
    box-shadow: 0 4px 8px rgba(5,150,105,0.25) !important;
    transition: all 0.15s ease !important;
}
button[kind="primary"] p,
button[kind="primary"] span,
button[kind="primaryFormSubmit"] p,
button[kind="primaryFormSubmit"] span,
button[data-testid="baseButton-primary"] p,
button[data-testid="baseButton-primary"] span,
div[data-testid="stFormSubmitButton"] > button p,
div[data-testid="stFormSubmitButton"] > button span,
div[data-testid="stFormSubmitButton"] > button * {
    color: #FFFFFF !important;
}
button[kind="primary"]:hover,
button[kind="primaryFormSubmit"]:hover,
button[data-testid="baseButton-primary"]:hover,
div[data-testid="stFormSubmitButton"] > button:hover {
    background: linear-gradient(to bottom, #10B981, #059669) !important;
    border-bottom-color: #047857 !important;
    box-shadow: 0 6px 16px rgba(16,185,129,0.35) !important;
    transform: translateY(-2px) !important;
}
button[kind="primary"]:disabled,
button[kind="primaryFormSubmit"]:disabled,
button[data-testid="baseButton-primary"]:disabled,
div[data-testid="stFormSubmitButton"] > button:disabled {
    background: #E2E8F0 !important;
    color: #94A3B8 !important;
    border-color: #CBD5E1 !important;
    border-bottom: 1px solid #CBD5E1 !important;
    box-shadow: none !important;
    cursor: not-allowed !important;
    transform: none !important;
}

/* ─── MODAL DIALOG PRIMARY BUTTON FIX ───────────────────────────────── */
div[data-testid="stModal"] button[data-testid="baseButton-primary"],
div[data-testid="stModal"] button[kind="primary"] {
    background: linear-gradient(to bottom, #059669, #047857) !important;
    color: #FFFFFF !important;
    border-left: 1px solid #065F46 !important;
    border-right: 1px solid #065F46 !important;
    border-top: 1px solid #065F46 !important;
    border-bottom: 4.5px solid #064E3B !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    box-shadow: 0 4px 8px rgba(5,150,105,0.25) !important;
}
div[data-testid="stModal"] button[data-testid="baseButton-primary"]:hover,
div[data-testid="stModal"] button[kind="primary"]:hover {
    background: linear-gradient(to bottom, #10B981, #059669) !important;
    border-bottom-color: #047857 !important;
    transform: translateY(-2px) !important;
}

/* ─── HIDE STREAMLIT TOOLBAR ─────────────────────────────────────────── */
header[data-testid="stHeader"] { display: none; }
div[data-testid="stToolbar"] { display: none; }
button[kind="header"] { display: none !important; }
</style>
""", unsafe_allow_html=True)

# -----------------------------
# HERO LAYOUT HEADER
# -----------------------------
st.markdown("""
<div class="hero-panel">
    <div class="hero-glow"></div>
    <div style="position: relative; z-index: 2; display: flex; align-items: center; gap: 16px;">
        <div class="hero-icon">💼</div>
        <div>
            <h1 class="hero-title">AI Powered Double-Entry Ledger</h1>
            <div class="hero-badge">Double Entry Active</div>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

# -----------------------------
# PERSISTENT NOTIFICATION TOAST/ALERT
# -----------------------------
notification_spot = st.empty()
if "success_msg" in st.session_state and st.session_state.success_msg:
    notification_spot.success(st.session_state.success_msg)
    del st.session_state.success_msg

# -----------------------------
# INTEREST POSTING REMINDER
# -----------------------------
missing_interest_months = []
actual_months = set()
for jv in get_all_journal_entries():
    if jv.get("jv_id") == "JV-VAL-GOLD":
        continue
    jv_date = to_date_obj(jv["date"])
    if jv_date:
        actual_months.add(jv_date.strftime("%Y-%m"))

for m_key in sorted(list(actual_months)):
    has_interest = False
    for jv in get_all_journal_entries():
        if jv.get("jv_id") == "JV-VAL-GOLD":
            continue
        jv_date = to_date_obj(jv["date"])
        if jv_date and jv_date.strftime("%Y-%m") == m_key:
            # Check narration
            if "interest" in jv.get("narration", "").lower():
                has_interest = True
                break
            # Check lines
            for line in jv.get("lines", []):
                if "interest" in line.get("account", "").lower():
                    has_interest = True
                    break
            if has_interest:
                break
    if not has_interest:
        try:
            dt_obj = datetime.strptime(m_key, "%Y-%m")
            formatted_m = dt_obj.strftime("%B %Y")
        except Exception:
            formatted_m = m_key
        missing_interest_months.append(formatted_m)

interest_reminder_spot = st.empty()
if missing_interest_months:
    months_str = ", ".join(missing_interest_months)
    interest_reminder_spot.markdown(f"""
    <div style="background: #FFFBEB; border: 1px solid #FEF3C7; border-left: 5px solid #F59E0B; border-radius: 12px; padding: 14px 20px; margin-bottom: 1.5rem; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.05);">
        <div style="display: flex; align-items: center; gap: 12px;">
            <span style="font-size: 1.25rem;">📢</span>
            <div style="font-family: 'Plus Jakarta Sans', sans-serif;">
                <strong style="color: #92400E; font-size: 0.95rem; font-weight: 700;">Interest Posting Reminder</strong>
                <p style="margin: 2px 0 0 0; font-size: 0.85rem; color: #B45309; line-height: 1.4;">
                    No interest transactions have been recorded for: <strong>{months_str}</strong>. Please post interest entries for these months.
                </p>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)



# -----------------------------
# SIDEBAR LEDGER SETTINGS
# -----------------------------
st.sidebar.header("Workspace Settings")
owner_name_input = st.sidebar.text_input("Change Owner Name", value=st.session_state.get("owner_name", "Abhilash"), key="owner_name_input")
if owner_name_input != st.session_state.get("owner_name", "Abhilash"):
    st.session_state.owner_name = owner_name_input
    save_data()
    st.rerun()

with st.sidebar.expander("☁️ Cloud Sync (Mobile Link)", expanded=False):
    st.markdown("**Real-time Cloud Sync**")
    sync_enabled = st.checkbox("Enable Cloud Sync", value=st.session_state.get("cloud_sync_enabled", False), key="sync_enabled_chk")
    cloud_url_in = st.text_input("Firebase RTDB URL", value=st.session_state.get("cloud_url", ""), placeholder="https://<proj>-default-rtdb.firebaseio.com", key="cloud_url_in")
    cloud_sec_in = st.text_input("Database Secret / Auth Token", value=st.session_state.get("cloud_secret", ""), type="password", placeholder="Enter database secret...", key="cloud_sec_in")
    
    if st.button("Update Cloud Settings", use_container_width=True, key="btn_update_cloud"):
        st.session_state.cloud_sync_enabled = sync_enabled
        st.session_state.cloud_url = cloud_url_in.strip()
        st.session_state.cloud_secret = cloud_sec_in.strip()
        save_data()
        st.session_state.success_msg = "✅ Cloud settings updated and synced!"
        st.rerun()

with st.sidebar.expander("⚙️ Security Credentials", expanded=False):
    st.markdown("**Change Password**")
    old_pw = st.text_input("Current Password", type="password", key="sb_old_pw")
    new_pw = st.text_input("New Password", type="password", key="sb_new_pw")
    confirm_pw = st.text_input("Confirm New Password", type="password", key="sb_confirm_pw")
    
    if st.button("Update Password", use_container_width=True, key="btn_sb_pw_update"):
        correct_pw = st.session_state.get("app_password", "finance@2026")
        if old_pw != correct_pw:
            st.error("Incorrect current password.")
        elif not new_pw:
            st.error("Password cannot be empty.")
        elif new_pw != confirm_pw:
            st.error("New passwords do not match.")
        else:
            st.session_state.app_password = new_pw
            save_data()
            
            # Clear text fields in session state
            st.session_state.sb_old_pw = ""
            st.session_state.sb_new_pw = ""
            st.session_state.sb_confirm_pw = ""
            st.session_state.success_msg = "✅ Password updated successfully!"
            st.rerun()
            
    st.markdown("---")
    st.markdown("**Reset Question**")
    new_q = st.text_input("Security Question", value=st.session_state.get("security_question", "What is the owner name of this finance ledger?"), key="sb_new_q")
    new_a = st.text_input("Security Answer", value=st.session_state.get("security_answer", "Abhilash"), key="sb_new_a")
    
    if st.button("Update Security Question", use_container_width=True, key="btn_sb_q_update"):
        if not new_q.strip() or not new_a.strip():
            st.error("Question and Answer cannot be empty.")
        else:
            st.session_state.security_question = new_q.strip()
            st.session_state.security_answer = new_a.strip()
            save_data()
            st.session_state.success_msg = "✅ Security question updated!"
            st.rerun()

# Logout button moved to main header

st.sidebar.header("Ledger Filters")
search_query = st.sidebar.text_input("Search Narration", "", placeholder="Search description...", key="sb_search")
filter_accs = st.sidebar.multiselect("Filter by Account", list(st.session_state.accounts.keys()), default=[])

# Date range selection
default_start = datetime(2026, 1, 1).date()
default_end = datetime(2027, 12, 31).date()
date_range = st.sidebar.date_input("Date Range", value=[default_start, default_end])

# Display Running Account Balances in Sidebar
st.sidebar.markdown("---")
st.sidebar.header("Account Balances")

liquid_bals = {}
invest_bals = {}
liab_bals = {}

for acc_name in st.session_state.accounts:
    acc_type = get_account_type(acc_name)
    bal = get_account_balance(acc_name)
    if acc_type == "Asset":
        if "bank" in acc_name.lower() or "cash" in acc_name.lower():
            liquid_bals[acc_name] = bal
        else:
            invest_bals[acc_name] = bal
    elif acc_type == "Liability":
        liab_bals[acc_name] = bal

st.sidebar.markdown("### 🏦 Cash & Bank")
for name, bal in liquid_bals.items():
    st.sidebar.markdown(f"**{name}**: ₹{bal:,.2f}")

if invest_bals:
    st.sidebar.markdown("### 📈 Investments")
    for name, bal in invest_bals.items():
        st.sidebar.markdown(f"**{name}**: ₹{bal:,.2f}")

if liab_bals:
    st.sidebar.markdown("### 💳 Liabilities & Loans")
    for name, bal in liab_bals.items():
        st.sidebar.markdown(f"**{name}**: ₹{bal:,.2f}")

# Apply pipeline filters on journal entries
filtered_jvs = []
for jv in get_all_journal_entries():
    if search_query.strip() and search_query.lower() not in jv["narration"].lower():
        continue
        
    jv_accs = [l["account"] for l in jv["lines"]]
    if filter_accs and not any(acc in jv_accs for acc in filter_accs):
        continue
        
    if isinstance(date_range, (list, tuple)) and len(date_range) == 2:
        start_d, end_d = date_range
        if not (start_d <= jv["date"] <= end_d):
            continue
            
    filtered_jvs.append(jv)

# -----------------------------
# TRANSACTION POSTING LAYOUT (QUICK & ADVANCED)
# -----------------------------
st.markdown("""
<div style="display:flex; align-items:center; gap:10px; margin-bottom:0.6rem;">
    <div style="width:36px; height:36px; border-radius:10px; background:linear-gradient(135deg,#10B981,#059669); display:flex; align-items:center; justify-content:center; font-size:1.1rem; box-shadow:0 4px 10px rgba(16,185,129,0.3);">➕</div>
    <div>
        <div style="font-family:'Outfit',sans-serif;font-size:1.1rem;font-weight:700;color:#0F172A;">Quick Add Transaction</div>
        <div style="font-size:0.8rem;color:#64748B;font-weight:500;">Fill standard fields to instantly log a double-entry transaction.</div>
    </div>
</div>
""", unsafe_allow_html=True)
with st.container():
    qe_c1, qe_c2, qe_c3, qe_c4, qe_c5 = st.columns([1.5, 2.2, 1.8, 1.8, 1.2])
    
    v = st.session_state.qe_form_version
    with qe_c1:
        qe_date = st.date_input("📅 Date", value=datetime.now().date(), key=f"qe_date_{v}")
        
    with qe_c2:
        if "quick_narration_val" not in st.session_state:
            st.session_state.quick_narration_val = ""
            
        qe_narration = st.text_input(
            "Narration",
            value=st.session_state.quick_narration_val,
            placeholder="e.g. Spent for groceries, utility...",
            key=f"qe_narration_input_{v}"
        )
        st.session_state.quick_narration_val = qe_narration
            
    with qe_c3:
        debit_options = [None] + list(st.session_state.accounts.keys())
        if "qe_debit_acc_val" not in st.session_state:
            st.session_state.qe_debit_acc_val = None
        sel_dr_idx = debit_options.index(st.session_state.qe_debit_acc_val) if st.session_state.qe_debit_acc_val in debit_options else 0
        
        qe_debit_acc = st.selectbox(
            "📥 Debit Account (Paid to)",
            options=debit_options,
            index=sel_dr_idx,
            format_func=format_account_label,
            key=f"qe_debit_acc_selectbox_{v}"
        )
        st.session_state.qe_debit_acc_val = qe_debit_acc
        
    with qe_c4:
        credit_options = [None] + list(st.session_state.accounts.keys())
        if "qe_credit_acc_val" not in st.session_state:
            st.session_state.qe_credit_acc_val = None
        sel_cr_idx = credit_options.index(st.session_state.qe_credit_acc_val) if st.session_state.qe_credit_acc_val in credit_options else 0
        
        qe_credit_acc = st.selectbox(
            "📤 Credit Account (Paid from)",
            options=credit_options,
            index=sel_cr_idx,
            format_func=format_account_label,
            key=f"qe_credit_acc_selectbox_{v}"
        )
        st.session_state.qe_credit_acc_val = qe_credit_acc
        
    with qe_c5:
        if "qe_amount_val" not in st.session_state:
            st.session_state.qe_amount_val = None
        val_to_use = float(st.session_state.qe_amount_val) if st.session_state.qe_amount_val is not None else None
        qe_amount = st.number_input(
            "Amount (INR)",
            min_value=0.01,
            value=val_to_use,
            step=100.0,
            format="%.2f",
            placeholder="0.00",
            key=f"qe_amount_input_{v}"
        )
        st.session_state.qe_amount_val = qe_amount

    # Low priority suggestions dropdown
    unique_narrations = []
    for jv in get_all_journal_entries():
        nar = jv.get("narration", "").strip()
        if nar and nar not in unique_narrations and not nar.startswith("Dynamic Gold"):
            unique_narrations.append(nar)
    unique_narrations = sorted(unique_narrations)

    suggestion_key = f"quick_narration_suggestion_select_{v}"
    selected_quick_narr = st.selectbox(
        "💡 Reuse Past Narration (Optional)",
        options=["-- Select to reuse narration --"] + unique_narrations,
        key=suggestion_key,
        on_change=on_quick_narration_change,
        args=(suggestion_key,)
    )

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
    qe_submit = st.button("Post Entry", use_container_width=True, type="primary", key="btn_quick_post")

    if qe_submit:
        if not qe_narration.strip():
            st.error("⚠️ Narration is required.")
        elif not qe_credit_acc or not qe_debit_acc:
            st.error("⚠️ Please select both Credit and Debit accounts.")
        elif qe_debit_acc == qe_credit_acc:
            st.error("⚠️ Debit and Credit accounts must be different.")
        elif not qe_amount or qe_amount <= 0:
            st.error("⚠️ Enter a valid amount greater than zero.")
        else:
            next_id = f"JV-{len(st.session_state.journal_entries) + 1:05d}"
            lines = [
                {"account": qe_debit_acc, "debit": float(qe_amount), "credit": 0.0},
                {"account": qe_credit_acc, "debit": 0.0, "credit": float(qe_amount)}
            ]
            st.session_state.journal_entries.append({
                "jv_id": next_id,
                "date": qe_date,
                "narration": qe_narration.strip(),
                "lines": lines
            })
            
            # Reset quick add form states
            st.session_state.quick_narration_val = ""
            st.session_state.quick_narration_sel = "Custom Narration..."
            st.session_state.qe_amount_val = None
            st.session_state.qe_credit_acc_val = None
            st.session_state.qe_debit_acc_val = None
            
            # Increment the form version to force a complete recreation of all widgets
            st.session_state.qe_form_version += 1
            
            save_data()
            st.session_state.success_msg = f"✅ Posted Transaction {next_id} successfully!"
            st.rerun()

st.markdown("---")

# -----------------------------
# ADVANCED JOURNAL VOUCHER SPLITS EDITOR
# -----------------------------
with st.expander("🛠️ Advanced Journal Voucher (JV) Splits Editor", expanded=False):
    st.markdown("""
    <div style="font-size:0.85rem;color:#64748B;font-weight:500;margin-bottom:0.8rem;">
        Log a complex multi-legged financial split (e.g. single payment split among tax, utilities, and assets).
        Ensure total Debits balance total Credits before posting.
    </div>
    """, unsafe_allow_html=True)
    
    v_adv = st.session_state.adv_form_version
    jv_date = st.date_input("JV Date", value=datetime.now().date(), key=f"jv_new_date_{v_adv}")
    if "adv_narration_val" not in st.session_state:
        st.session_state.adv_narration_val = ""
        
    jv_narration = st.text_input(
        "Narration / Description",
        value=st.session_state.adv_narration_val,
        placeholder="e.g. Salary distribution, split business expenses...",
        key=f"jv_new_narration_input_{v_adv}"
    )
    st.session_state.adv_narration_val = jv_narration
        
    # Low priority suggestions dropdown
    unique_narrations = []
    for jv in get_all_journal_entries():
        nar = jv.get("narration", "").strip()
        if nar and nar not in unique_narrations and not nar.startswith("Dynamic Gold"):
            unique_narrations.append(nar)
    unique_narrations = sorted(unique_narrations)
    
    adv_suggestion_key = f"jv_narration_suggestion_select_{v_adv}"
    selected_narr_suggestion = st.selectbox(
        "💡 Reuse Past Narration (Optional)",
        options=["-- Select to reuse narration --"] + unique_narrations,
        key=adv_suggestion_key,
        on_change=on_adv_narration_change,
        args=(adv_suggestion_key,)
    )

    # Initialize splits list if not present
    if "new_jv_lines_list" not in st.session_state:
        st.session_state.new_jv_lines_list = [
            {"id": "row_1", "account": None, "debit": 0.0, "credit": 0.0},
            {"id": "row_2", "account": None, "debit": 0.0, "credit": 0.0}
        ]

    st.markdown("##### Splits Distribution")
    
    # Column headers
    th_col1, th_col2, th_col3, th_col4 = st.columns([3, 2, 2, 0.5])
    th_col1.markdown("**Account**")
    th_col2.markdown("**Debit (INR)**")
    th_col3.markdown("**Credit (INR)**")
    th_col4.write("")

    has_empty_account = False
    lines_to_save = []
    
    # Display each line as standard inputs
    for idx in range(len(st.session_state.new_jv_lines_list)):
        line = st.session_state.new_jv_lines_list[idx]
        col_acc, col_dr, col_cr, col_del = st.columns([3, 2, 2, 0.5])
        
        with col_acc:
            account_options = [None] + list(st.session_state.accounts.keys())
            current_acc = line["account"]
            if current_acc not in account_options:
                current_acc = None
            sel_idx = account_options.index(current_acc)
            selected_acc = st.selectbox(
                f"Account_{idx}",
                options=account_options,
                index=sel_idx,
                format_func=format_account_label,
                key=f"adv_jv_acc_{v_adv}_{line['id']}",
                label_visibility="collapsed"
            )
            line["account"] = selected_acc
            
        with col_dr:
            debit_val = st.number_input(
                f"Debit_{idx}",
                min_value=0.0,
                value=float(line["debit"]),
                step=100.0,
                format="%.2f",
                key=f"adv_jv_dr_{v_adv}_{line['id']}",
                label_visibility="collapsed"
            )
            line["debit"] = debit_val
            
        with col_cr:
            credit_val = st.number_input(
                f"Credit_{idx}",
                min_value=0.0,
                value=float(line["credit"]),
                step=100.0,
                format="%.2f",
                key=f"adv_jv_cr_{v_adv}_{line['id']}",
                label_visibility="collapsed"
            )
            line["credit"] = credit_val
            
        with col_del:
            if st.button("🗑️", key=f"adv_jv_del_{line['id']}", help="Delete this row"):
                st.session_state.new_jv_lines_list.pop(idx)
                st.rerun()
                
        if (debit_val > 0 or credit_val > 0) and selected_acc is None:
            has_empty_account = True
            
        if selected_acc is not None and (debit_val > 0 or credit_val > 0):
            lines_to_save.append({
                "account": selected_acc,
                "debit": debit_val,
                "credit": credit_val
            })
            
    # Add Row Button
    if st.button("➕ Add Line", key="btn_add_line_adv"):
        st.session_state.new_jv_lines_list.append({
            "id": f"row_{datetime.now().timestamp()}_{len(st.session_state.new_jv_lines_list)}",
            "account": None,
            "debit": 0.0,
            "credit": 0.0
        })
        st.rerun()

    # Calculate live balance check
    total_debit = sum(l["debit"] for l in st.session_state.new_jv_lines_list)
    total_credit = sum(l["credit"] for l in st.session_state.new_jv_lines_list)
    balance_diff = total_debit - total_credit

    st.markdown(f"**Total Debits**: ₹{total_debit:,.2f}  |  **Total Credits**: ₹{total_credit:,.2f}")
    
    is_balanced = abs(balance_diff) < 0.01 and total_debit > 0
    if is_balanced:
        st.success("✅ JV is balanced perfectly and ready to post.")
        
        if has_empty_account:
            st.warning("⚠️ One or more split lines have an amount but no account selected. Please select an account or remove the line.")
            
        if st.button("💾 Post Journal Voucher", type="primary", use_container_width=True, disabled=has_empty_account, key="btn_post_jv"):
            next_id = f"JV-{len(st.session_state.journal_entries) + 1:05d}"
            
            st.session_state.journal_entries.append({
                "jv_id": next_id,
                "date": jv_date,
                "narration": jv_narration.strip() if jv_narration.strip() else "Manual JV Splits",
                "lines": lines_to_save
            })
            save_data()
            st.session_state.success_msg = f"✅ Journal Voucher {next_id} posted successfully!"
            
            # Reset editor state to empty (None) with fresh unique IDs
            st.session_state.new_jv_lines_list = [
                {"id": f"row_{datetime.now().timestamp()}_1", "account": None, "debit": 0.0, "credit": 0.0},
                {"id": f"row_{datetime.now().timestamp()}_2", "account": None, "debit": 0.0, "credit": 0.0}
            ]
            st.session_state.adv_narration_val = ""
            st.session_state.adv_narration_sel = "Custom Narration..."
            
            # Increment version to force complete UI reset
            st.session_state.adv_form_version += 1
            st.rerun()
            
    elif total_debit == 0 and total_credit == 0:
        st.info("💡 Add debit and credit legs to the transaction lines above.")
    else:
        st.error(f"❌ Unbalanced Entry: Difference is ₹{abs(balance_diff):,.2f}. Debits and Credits must balance to zero before posting.")

# -----------------------------
# LIVE SUMMARY METRICS (BALANCE SHEET VIEW)
# -----------------------------
# Calculate core values
total_assets = 0.0
total_liabilities = 0.0
total_equity_accts = 0.0
total_revenue = 0.0
total_expense = 0.0

for acc_name in st.session_state.accounts:
    bal = get_account_balance(acc_name)
    acc_type = get_account_type(acc_name)
    if acc_type == "Asset":
        total_assets += bal
    elif acc_type == "Liability":
        total_liabilities += bal
    elif acc_type == "Equity":
        total_equity_accts += bal
    elif acc_type == "Revenue":
        total_revenue += bal
    elif acc_type == "Expense":
        total_expense += bal

net_profit = total_revenue - total_expense
total_equity = total_equity_accts + net_profit

# Calculate Cash & Bank Balance
liquid_assets_total = 0.0
for acc_name in st.session_state.accounts:
    acc_type = get_account_type(acc_name)
    if acc_type == "Asset" and ("bank" in acc_name.lower() or "cash" in acc_name.lower()):
        liquid_assets_total += get_account_balance(acc_name)

# Calculate total loan payments
total_loan_payments = 0.0
for jv in get_all_journal_entries():
    for line in jv["lines"]:
        acc_name = line["account"]
        acc_type = get_account_type(acc_name)
        if acc_type == "Liability":
            total_loan_payments += line["debit"]

net_savings_val = total_assets
net_savings_delta = total_revenue - total_expense - total_loan_payments

# Calculate Receipts and Payout dynamically from cash/bank inflows and outflows
total_receipts = 0.0
total_payout = 0.0
payout_breakdown = {}

for jv in get_all_journal_entries():
    # Find cash/bank lines in this JV
    cash_bank_credits = []
    cash_bank_debits = []
    other_debits = []
    
    for line in jv["lines"]:
        acc_name = line["account"]
        acc_type = get_account_type(acc_name)
        is_cash_bank = acc_type == "Asset" and ("bank" in acc_name.lower() or "cash" in acc_name.lower())
        
        if is_cash_bank:
            if line["credit"] > 0:
                cash_bank_credits.append(line)
            if line["debit"] > 0:
                cash_bank_debits.append(line)
        else:
            if line["debit"] > 0:
                other_debits.append(line)
                
    # 1. Process Outflows (Payouts)
    if cash_bank_credits:
        # Check if it's an internal transfer (debit is also a cash/bank account)
        is_internal = len(cash_bank_debits) > 0
        if not is_internal:
            for line in cash_bank_credits:
                total_payout += line["credit"]
            for od in other_debits:
                od_acc = od["account"]
                od_type = get_account_type(od_acc)
                amount = od["debit"]
                
                if od_type == "Expense":
                    cat = "Expense"
                elif "BlinkX" in od_acc or "Blinx" in od_acc:
                    cat = "Transfer to BlinkX"
                elif "Bajaj" in od_acc:
                    cat = "Transfer to Bajaj Finance"
                elif od_type == "Asset":
                    disp = od_acc.replace(" Account", "").replace(" Asset", "")
                    cat = f"Transfer to {disp}"
                elif od_type == "Liability":
                    disp = od_acc.replace(" Loan", "")
                    cat = f"Transfer to {disp}"
                else:
                    cat = od_acc
                payout_breakdown[cat] = payout_breakdown.get(cat, 0.0) + amount
                
    # 2. Process Inflows (Receipts)
    if cash_bank_debits:
        # Check if it's an internal transfer (credit is also a cash/bank account)
        is_internal = len(cash_bank_credits) > 0
        if not is_internal:
            for line in cash_bank_debits:
                total_receipts += line["debit"]

col_m1, col_m2, col_m3 = st.columns(3)
with col_m1:
    st.metric("Net Receipts", f"₹{total_receipts:,.2f}")

with col_m2:
    st.metric("Net Payout", f"₹{total_payout:,.2f}")
    if payout_breakdown:
        breakdown_items = []
        for cat, amt in payout_breakdown.items():
            if amt > 0:
                breakdown_items.append(f'<div style="display: flex; justify-content: space-between; font-size: 0.8rem; line-height: 1.4; border-bottom: 1px dashed #F1F5F9; padding: 2px 0;"><span style="font-weight: 600; color: #475569;">📍 {cat}</span><span style="font-weight: 700; color: #0F172A;">₹{amt:,.2f}</span></div>')
        
        st.markdown(f'<div style="background: #FFFFFF; border: 1px solid #E2E8F0; border-radius: 12px; padding: 10px 14px; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.03); margin-top: 8px;"><div style="font-family: Outfit, sans-serif; font-size: 0.68rem; color: #94A3B8; font-weight: 700; text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 6px;">📋 Payout Breakdown</div>{"".join(breakdown_items)}</div>', unsafe_allow_html=True)

with col_m3:
    st.metric("Net Balance", f"₹{liquid_assets_total:,.2f}")

st.markdown("---")

# -----------------------------
# DASHBOARD TABS ENGINE
# -----------------------------
tab_jl, tab_gl, tab_tb, tab_pl, tab_bs, tab_coa, tab_fc, tab_an, tab_io, tab_ev = st.tabs([
    "📋 Journal Ledger",
    "📖 General Ledger",
    "📊 Trial Balance",
    "📈 Profit & Loss",
    "🏦 Balance Sheet",
    "⚙️ Chart of Accounts",
    "🔮 Asset Projections",
    "📊 Analytics",
    "📥 Template Sync",
    "🚗 Important Events"
], key="main_navigation_tabs")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 — JOURNAL LEDGER
# ─────────────────────────────────────────────────────────────────────────────
with tab_jl:
    # 1. Inline Editor (if edit mode active)
    if "edit_mode" not in st.session_state:
        st.session_state.edit_mode = False
        st.session_state.edit_target = None
        
    if st.session_state.edit_mode and st.session_state.edit_target:
        target_jv_id = st.session_state.edit_target
        target_jv = next((jv for jv in st.session_state.journal_entries if jv["jv_id"] == target_jv_id), None)
        if target_jv:
            st.markdown(f"""
            <div style="background-color: #EFF6FF; border: 1.5px solid #BFDBFE; border-radius: 12px; padding: 18px; margin-bottom: 20px; box-shadow: 0 4px 12px rgba(59, 130, 246, 0.05);">
                <span style="font-size: 0.85rem; color: #1D4ED8; font-weight: 800; text-transform: uppercase; letter-spacing: 0.05em;">📝 EDIT JOURNAL ENTRY</span>
                <h4 style="margin: 4px 0 10px 0; color: #1E3A8A; font-weight: 700;">Journal Voucher: {target_jv_id}</h4>
            </div>
            """, unsafe_allow_html=True)
            
            ec1, ec2 = st.columns([1, 2])
            with ec1:
                new_date = st.date_input("📅 Date", value=target_jv["date"], key=f"inline_edit_date_{target_jv_id}")
            with ec2:
                # Initialize state for edit narration value
                state_narr_key = f"edit_narration_val_{target_jv_id}"
                if state_narr_key not in st.session_state:
                    st.session_state[state_narr_key] = target_jv["narration"]
                
                new_narration = st.text_input(
                    "Narration",
                    value=st.session_state[state_narr_key],
                    key=f"inline_edit_narr_{target_jv_id}"
                )
                st.session_state[state_narr_key] = new_narration
                
                # Narration selector suggestions for edit mode
                unique_narrations = []
                for jv in get_all_journal_entries():
                    nar = jv.get("narration", "").strip()
                    if nar and nar not in unique_narrations and not nar.startswith("Dynamic Gold"):
                        unique_narrations.append(nar)
                unique_narrations = sorted(unique_narrations)
                
                edit_suggestion_key = f"edit_narration_suggestion_select_{target_jv_id}"
                selected_suggestion = st.selectbox(
                    "💡 Reuse Past Narration (Optional)",
                    options=["-- Select to reuse narration --"] + unique_narrations,
                    key=edit_suggestion_key,
                    on_change=on_edit_narration_change,
                    args=(state_narr_key, edit_suggestion_key)
                )

                
            st.markdown("##### Splits Distribution")
            
            # Retrieve splits from state or init
            state_key = f"editing_jv_lines_{target_jv_id}"
            if state_key not in st.session_state:
                st.session_state[state_key] = [
                    {
                        "id": f"row_{idx}",
                        "account": line["account"],
                        "debit": float(line["debit"]),
                        "credit": float(line["credit"])
                    }
                    for idx, line in enumerate(target_jv["lines"])
                ]
                
            account_options = list(st.session_state.accounts.keys())
            for line in target_jv["lines"]:
                if line["account"] not in account_options:
                    account_options.append(line["account"])
                    
            # Render headers
            col_eh1, col_eh2, col_eh3, col_eh4 = st.columns([3, 2, 2, 0.5])
            col_eh1.markdown("**Account**")
            col_eh2.markdown("**Debit (INR)**")
            col_eh3.markdown("**Credit (INR)**")
            col_eh4.write("")
            
            has_empty_account = False
            lines_to_save_inline = []
            
            # We display each line
            for idx in range(len(st.session_state[state_key])):
                line = st.session_state[state_key][idx]
                col_acc, col_dr, col_cr, col_del = st.columns([3, 2, 2, 0.5])
                
                with col_acc:
                    options = [None] + account_options
                    current_acc = line["account"]
                    if current_acc not in options:
                        options.append(current_acc)
                    sel_idx = options.index(current_acc)
                    selected_acc = st.selectbox(
                        f"Edit_Account_{idx}",
                        options=options,
                        index=sel_idx,
                        format_func=format_account_label,
                        key=f"edit_jv_acc_{target_jv_id}_{line['id']}",
                        label_visibility="collapsed"
                    )
                    line["account"] = selected_acc
                    
                with col_dr:
                    debit_val = st.number_input(
                        f"Edit_Debit_{idx}",
                        min_value=0.0,
                        value=float(line["debit"]),
                        step=100.0,
                        format="%.2f",
                        key=f"edit_jv_dr_{target_jv_id}_{line['id']}",
                        label_visibility="collapsed"
                    )
                    line["debit"] = debit_val
                    
                with col_cr:
                    credit_val = st.number_input(
                        f"Edit_Credit_{idx}",
                        min_value=0.0,
                        value=float(line["credit"]),
                        step=100.0,
                        format="%.2f",
                        key=f"edit_jv_cr_{target_jv_id}_{line['id']}",
                        label_visibility="collapsed"
                    )
                    line["credit"] = credit_val
                    
                with col_del:
                    if st.button("🗑️", key=f"edit_jv_del_{target_jv_id}_{line['id']}", help="Delete this row"):
                        st.session_state[state_key].pop(idx)
                        st.rerun()
                        
                if (debit_val > 0 or credit_val > 0) and selected_acc is None:
                    has_empty_account = True
                    
                if selected_acc is not None and (debit_val > 0 or credit_val > 0):
                    lines_to_save_inline.append({
                        "account": selected_acc,
                        "debit": debit_val,
                        "credit": credit_val
                    })
                    
            if st.button("➕ Add Line", key=f"btn_add_line_edit_{target_jv_id}"):
                st.session_state[state_key].append({
                    "id": f"row_{datetime.now().timestamp()}_{len(st.session_state[state_key])}",
                    "account": None,
                    "debit": 0.0,
                    "credit": 0.0
                })
                st.rerun()
                
            # Calculate live balance check
            total_debit = sum(l["debit"] for l in st.session_state[state_key])
            total_credit = sum(l["credit"] for l in st.session_state[state_key])
            balance_diff = total_debit - total_credit
            
            st.markdown(f"**Total Debits**: ₹{total_debit:,.2f}  |  **Total Credits**: ₹{total_credit:,.2f}")
            
            is_balanced = abs(balance_diff) < 0.01 and total_debit > 0
            if is_balanced:
                st.success("✅ JV is balanced perfectly.")
            else:
                st.error(f"❌ Unbalanced Entry: Difference is ₹{abs(balance_diff):,.2f}. Debits and Credits must balance to zero.")

            eb1, eb2 = st.columns(2)
            with eb1:
                # Save button (only clickable if balanced and has no empty accounts)
                if st.button("💾 Save Changes", type="primary", use_container_width=True, disabled=(not is_balanced or has_empty_account), key=f"btn_save_inline_{target_jv_id}"):
                    target_jv["date"] = new_date
                    target_jv["narration"] = new_narration.strip()
                    target_jv["lines"] = lines_to_save_inline
                    save_data()
                    
                    # Reset edit mode
                    st.session_state.edit_mode = False
                    st.session_state.edit_target = None
                    
                    # Delete all session state keys associated with editing this JV
                    for k in list(st.session_state.keys()):
                        if target_jv_id in k or k == state_key:
                            del st.session_state[k]
                        
                    st.session_state.success_msg = "✅ Updated JV successfully!"
                    st.rerun()
                        
            with eb2:
                if st.button("❌ Cancel Editing", use_container_width=True, key=f"btn_cancel_inline_{target_jv_id}"):
                    st.session_state.edit_mode = False
                    st.session_state.edit_target = None
                    
                    # Delete all session state keys associated with editing this JV
                    for k in list(st.session_state.keys()):
                        if target_jv_id in k or k == state_key:
                            del st.session_state[k]
                    st.rerun()
            st.markdown("---")

    # 2. Main Tab layout (Logs list and controls)
    col_jl1, col_jl2 = st.columns([3, 1])
    with col_jl1:
        st.subheader("Journal Entry Logs")
    with col_jl2:
        jv_ids = [jv["jv_id"] for jv in st.session_state.journal_entries]
        if jv_ids:
            sorted_jv_ids = sorted(jv_ids, reverse=True)
            target_mgmt_jv = st.selectbox("Select JV to Manage", options=sorted_jv_ids, key="sel_mgmt_jv")
            
            if st.button("📝 Edit Selected JV", use_container_width=True, key="btn_edit_jv_trigger"):
                st.session_state.edit_mode = True
                st.session_state.edit_target = target_mgmt_jv
                state_key = f"editing_jv_lines_{target_mgmt_jv}"
                
                # Clean up any old keys for this target JV to be fresh
                for k in list(st.session_state.keys()):
                    if target_mgmt_jv in k or k == state_key:
                        del st.session_state[k]
                st.rerun()
                
            if st.button("🗑️ Delete Selected JV", use_container_width=True, key="btn_del_jv"):
                st.session_state.journal_entries = [jv for jv in st.session_state.journal_entries if jv["jv_id"] != target_mgmt_jv]
                save_data()
                st.session_state.success_msg = f"🗑️ Journal Entry {target_mgmt_jv} deleted successfully!"
                st.rerun()

    # Formulate a flat grid view of the journal entries
    jv_grid_rows = []
    for jv in filtered_jvs:
        for l in jv["lines"]:
            jv_grid_rows.append({
                "Date": jv["date"],
                "JV ID": jv["jv_id"],
                "Narration": jv["narration"],
                "Account": l["account"],
                "Debit": l["debit"],
                "Credit": l["credit"]
            })
            
    df_jv_grid = pd.DataFrame(jv_grid_rows)
    if not df_jv_grid.empty:
        # Push dynamic auto JVs (such as gold valuation JVs starting with JV-VAL-) to the bottom of the list
        df_jv_grid["IsAutoJV"] = df_jv_grid["JV ID"].astype(str).str.startswith("JV-VAL-")
        df_sorted = df_jv_grid.sort_values(
            by=["IsAutoJV", "Date", "JV ID"],
            ascending=[True, False, False]
        ).drop(columns=["IsAutoJV"])
        st.dataframe(df_sorted, use_container_width=True, height=400, hide_index=True)
    else:
        st.info("No matching journal entries found in transaction ledger.")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 — GENERAL LEDGER
# ─────────────────────────────────────────────────────────────────────────────
with tab_gl:
    st.subheader("General Ledger Book")
    gl_col1, gl_col2 = st.columns([2, 1])
    with gl_col1:
        sel_gl_account = st.selectbox("Select Account", options=list(st.session_state.accounts.keys()), key="sel_gl_account")
    with gl_col2:
        if sel_gl_account:
            final_bal = get_account_balance(sel_gl_account)
            st.metric("Final Balance (As on Date)", f"₹{final_bal:,.2f}")
    
    if sel_gl_account:
        df_gl = get_ledger_for_account(sel_gl_account)
        if not df_gl.empty:
            st.dataframe(
                df_gl.style.format({
                    "Debit": "₹{:.2f}",
                    "Credit": "₹{:.2f}",
                    "Running Balance": "₹{:.2f}"
                }),
                use_container_width=True,
                height=350,
                hide_index=True
            )
        else:
            st.info(f"No transactions posted to '{sel_gl_account}' yet.")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 — TRIAL BALANCE
# ─────────────────────────────────────────────────────────────────────────────
with tab_tb:
    st.subheader("Trial Balance")
    tb_level = st.radio("Report View Level", ["Detailed Sub-ledgers", "Main Ledgers Only"], key="tb_view_level", horizontal=True)
    
    trial_rows = []
    balances_to_use = get_rolled_up_balances(tb_level)
    
    for acc_name, info in balances_to_use.items():
        bal = info["balance"]
        acc_type = info["type"]
        if bal == 0.0:
            continue
            
        net_dr = 0.0
        net_cr = 0.0
        if acc_type in ["Asset", "Expense"]:
            if bal > 0:
                net_dr = bal
            else:
                net_cr = abs(bal)
        else:
            if bal > 0:
                net_cr = bal
            else:
                net_dr = abs(bal)
                
        trial_rows.append({
            "Account": acc_name,
            "Account Type": acc_type,
            "Debit Balance": net_dr,
            "Credit Balance": net_cr
        })

    df_trial = pd.DataFrame(trial_rows)
    if not df_trial.empty:
        st.dataframe(
            df_trial.style.format({
                "Debit Balance": "₹{:.2f}",
                "Credit Balance": "₹{:.2f}"
            }),
            use_container_width=True,
            hide_index=True
        )
        tot_dr_bal = df_trial["Debit Balance"].sum()
        tot_cr_bal = df_trial["Credit Balance"].sum()
        diff_bal = abs(tot_dr_bal - tot_cr_bal)
        
        st.markdown(f"**Total Debit Balance**: ₹{tot_dr_bal:,.2f}  |  **Total Credit Balance**: ₹{tot_cr_bal:,.2f}")
        if diff_bal < 0.01:
            st.success("✅ Trial Balance is perfectly in balance!")
        else:
            st.error(f"❌ Trial Balance is unbalanced by ₹{diff_bal:,.2f}. Check JVs for entry errors.")
    else:
        st.info("No active accounts found. Post entries to evaluate Trial Balance.")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 4 — PROFIT & LOSS
# ─────────────────────────────────────────────────────────────────────────────
with tab_pl:
    st.subheader("Profit & Loss Statement (Income Statement)")
    pl_level = st.radio("Report View Level", ["Detailed Sub-ledgers", "Main Ledgers Only"], key="pl_view_level", horizontal=True)
    
    balances_to_use = get_rolled_up_balances(pl_level)
    revenue_details = []
    expense_details = []
    total_rev_p = 0.0
    total_exp_p = 0.0
    
    for acc_name, info in balances_to_use.items():
        bal = info["balance"]
        acc_type = info["type"]
        if acc_type == "Revenue":
            total_rev_p += bal
            if bal != 0:
                revenue_details.append({"Revenue Account": acc_name, "Amount": bal})
        elif acc_type == "Expense":
            total_exp_p += bal
            if bal != 0:
                expense_details.append({"Expense Account": acc_name, "Amount": bal})

    col_pl1, col_pl2 = st.columns(2)
    
    with col_pl1:
        st.markdown("### 📥 Revenue")
        if revenue_details:
            df_rev = pd.DataFrame(revenue_details)
            st.dataframe(df_rev.style.format({"Amount": "₹{:.2f}"}), use_container_width=True, hide_index=True)
        else:
            st.caption("No revenue recorded.")
        st.markdown(f"**Total Revenue**: ₹{total_rev_p:,.2f}")

    with col_pl2:
        st.markdown("### 📤 Expenses")
        if expense_details:
            df_exp = pd.DataFrame(expense_details)
            st.dataframe(df_exp.style.format({"Amount": "₹{:.2f}"}), use_container_width=True, hide_index=True)
        else:
            st.caption("No expenses recorded.")
        st.markdown(f"**Total Expenses**: ₹{total_exp_p:,.2f}")

    st.markdown("---")
    net_profit_p = total_rev_p - total_exp_p
    if net_profit_p >= 0:
        st.success(f"### 🎉 Net Period Profit: ₹{net_profit_p:,.2f}")
    else:
        st.error(f"### 💸 Net Period Loss: ₹{abs(net_profit_p):,.2f}")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 5 — BALANCE SHEET
# ─────────────────────────────────────────────────────────────────────────────
with tab_bs:
    st.subheader("Balance Sheet")
    bs_level = st.radio("Report View Level", ["Detailed Sub-ledgers", "Main Ledgers Only"], key="bs_view_level", horizontal=True)
    
    balances_to_use = get_rolled_up_balances(bs_level)
    asset_details = []
    liability_details = []
    equity_details = []
    
    total_assets_b = 0.0
    total_liabilities_b = 0.0
    total_equity_accts_b = 0.0
    
    for acc_name, info in balances_to_use.items():
        bal = info["balance"]
        acc_type = info["type"]
        if acc_type == "Asset":
            total_assets_b += bal
            if bal != 0:
                asset_details.append({"Asset Account": acc_name, "Amount": bal})
        elif acc_type == "Liability":
            total_liabilities_b += bal
            if bal != 0:
                liability_details.append({"Liability Account": acc_name, "Amount": bal})
        elif acc_type == "Equity":
            total_equity_accts_b += bal
            if bal != 0:
                equity_details.append({"Equity Account": acc_name, "Amount": bal})
                
    # Calculate net profit from detailed for accuracy
    detailed_bals = get_rolled_up_balances("Detailed")
    total_rev_d = sum(info["balance"] for info in detailed_bals.values() if info["type"] == "Revenue")
    total_exp_d = sum(info["balance"] for info in detailed_bals.values() if info["type"] == "Expense")
    net_profit_d = total_rev_d - total_exp_d
    
    if net_profit_d != 0:
        equity_details.append({"Equity Account": "Current Period Net Income", "Amount": net_profit_d})

    col_bs1, col_bs2, col_bs3 = st.columns(3)
    
    with col_bs1:
        st.markdown("### 🏛️ Assets")
        if asset_details:
            df_assets = pd.DataFrame(asset_details)
            st.dataframe(df_assets.style.format({"Amount": "₹{:.2f}"}), use_container_width=True, hide_index=True)
        else:
            st.caption("No asset balances.")
        st.markdown(f"**Total Assets**: ₹{total_assets_b:,.2f}")

    with col_bs2:
        st.markdown("### 💳 Liabilities")
        if liability_details:
            df_liab = pd.DataFrame(liability_details)
            st.dataframe(df_liab.style.format({"Amount": "₹{:.2f}"}), use_container_width=True, hide_index=True)
        else:
            st.caption("No liability balances.")
        st.markdown(f"**Total Liabilities**: ₹{total_liabilities_b:,.2f}")

    with col_bs3:
        st.markdown("### 📈 Equity")
        if equity_details:
            df_eq = pd.DataFrame(equity_details)
            st.dataframe(df_eq.style.format({"Amount": "₹{:.2f}"}), use_container_width=True, hide_index=True)
        else:
            st.caption("No equity balances.")
        total_equity_b = total_equity_accts_b + net_profit_d
        st.markdown(f"**Total Equity**: ₹{total_equity_b:,.2f}")

    st.markdown("---")
    bs_sum = total_liabilities_b + total_equity_b
    col_eq1, col_eq2 = st.columns(2)
    col_eq1.markdown(f"#### Total Assets: ₹{total_assets_b:,.2f}")
    col_eq2.markdown(f"#### Total Liabilities + Equity: ₹{bs_sum:,.2f}")

    if abs(total_assets_b - bs_sum) < 0.01:
        st.success("✅ Balance Sheet matches perfectly! Assets = Liabilities + Equity")
    else:
        st.error(f"❌ Imbalance Detected: Difference is ₹{total_assets_b - bs_sum:,.2f}")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 6 — CHART OF ACCOUNTS
# ─────────────────────────────────────────────────────────────────────────────
with tab_coa:
    st.subheader("Chart of Accounts Manager")

    coa_col1, coa_col2, coa_col3 = st.columns(3)

    with coa_col1:
        with st.expander("➕ Add Account", expanded=False):
            new_acc_name = st.text_input("Account Name", key="txt_add_acc", placeholder="e.g. Credit Card PNB, Medical Expense")
            acc_type = st.selectbox(
                "Account Type",
                ["Asset", "Liability", "Equity", "Revenue", "Expense"],
                key="sel_acc_type",
                on_change=on_add_type_change
            )
            
            # Parent selection options (only accounts of same type that are top-level)
            parent_options = [None] + [name for name in st.session_state.accounts if get_account_type(name) == acc_type and get_account_parent(name) is None]
            new_acc_parent = st.selectbox("Parent Account (Optional)", parent_options, key="sel_add_acc_parent")
            
            if st.button("Create Account", use_container_width=True, key="btn_create_acc"):
                clean_name = new_acc_name.strip()
                if clean_name and clean_name not in st.session_state.accounts:
                    st.session_state.accounts[clean_name] = {"type": acc_type, "parent": new_acc_parent}
                    st.session_state.success_msg = f"✅ Added account '{clean_name}' ({acc_type}) successfully!"
                    save_data()
                    st.rerun()
                else:
                    st.error("Invalid name or account already exists.")

    with coa_col2:
        with st.expander("📝 Edit Account Settings", expanded=False):
            target_edit_acc = st.selectbox(
                "Select Account to Edit",
                list(st.session_state.accounts.keys()),
                key="sel_edit_target_acc",
                on_change=on_edit_target_change
            )
            new_edit_acc_name = st.text_input("Rename To", value=target_edit_acc, key="txt_edit_acc_name")
            current_type = get_account_type(target_edit_acc)
            type_options = ["Asset", "Liability", "Equity", "Revenue", "Expense"]
            new_edit_acc_type = st.selectbox(
                "Change Type To",
                options=type_options,
                index=type_options.index(current_type),
                key="sel_edit_acc_type",
                on_change=on_edit_type_change
            )
            
            # Parent selection (only same type, parent = None, not itself, and check if itself is a parent of others)
            has_children = any(get_account_parent(name) == target_edit_acc for name in st.session_state.accounts)
            if has_children:
                st.caption("ℹ️ This is a Parent Account. Set its children's parents to None before assigning a parent to this account.")
                edit_parent_options = [None]
            else:
                edit_parent_options = [None] + [name for name in st.session_state.accounts if get_account_type(name) == new_edit_acc_type and get_account_parent(name) is None and name != target_edit_acc]
            
            current_parent = get_account_parent(target_edit_acc)
            edit_parent_idx = edit_parent_options.index(current_parent) if current_parent in edit_parent_options else 0
            new_edit_parent = st.selectbox("Change Parent To", edit_parent_options, index=edit_parent_idx, key="sel_edit_acc_parent")
            
            if st.button("Save Changes", use_container_width=True, key="btn_rename_acc"):
                clean_name = new_edit_acc_name.strip()
                if clean_name:
                    if clean_name != target_edit_acc and clean_name in st.session_state.accounts:
                        st.error("An account with that name already exists.")
                    else:
                        st.session_state.accounts.pop(target_edit_acc, None)
                        st.session_state.accounts[clean_name] = {"type": new_edit_acc_type, "parent": new_edit_parent}
                        
                        # Update children referencing the old name
                        if clean_name != target_edit_acc:
                            for name, val in st.session_state.accounts.items():
                                if isinstance(val, dict) and val.get("parent") == target_edit_acc:
                                    val["parent"] = clean_name
                        
                        # Update all historical entries referencing this account
                        for jv in st.session_state.journal_entries:
                            for l in jv["lines"]:
                                if l["account"] == target_edit_acc:
                                    l["account"] = clean_name
                                    
                        st.session_state.success_msg = f"✅ Updated '{target_edit_acc}' settings!"
                        save_data()
                        st.rerun()

    with coa_col3:
        with st.expander("❌ Delete Account", expanded=False):
            target_remove_acc = st.selectbox("Select Account", list(st.session_state.accounts.keys()), key="sel_rem_target_acc")
            st.warning("⚠️ Deleting reassigns historical ledger transactions to 'Other Expenses'. Any children of this account will become top-level.")
            if st.button("Delete Account", use_container_width=True, key="btn_delete_acc"):
                if target_remove_acc in ["Cash", "Bank Account", "Other Expenses"]:
                    st.error("Cannot delete core system accounts.")
                else:
                    # Release children from deleted parent
                    for name, val in st.session_state.accounts.items():
                        if isinstance(val, dict) and val.get("parent") == target_remove_acc:
                            val["parent"] = None
                            
                    for jv in st.session_state.journal_entries:
                        for l in jv["lines"]:
                            if l["account"] == target_remove_acc:
                                l["account"] = "Other Expenses"
                                
                    st.session_state.accounts.pop(target_remove_acc, None)
                    st.session_state.success_msg = f"🗑️ Removed '{target_remove_acc}' completely."
                    save_data()
                    st.rerun()

    # Display accounts list organized by group type
    st.subheader("Active Accounts List")
    tree_rows = []
    for typ in ["Asset", "Liability", "Equity", "Revenue", "Expense"]:
        type_accs = [name for name in st.session_state.accounts if get_account_type(name) == typ]
        # Main ledgers (parent is None)
        parents = [name for name in type_accs if get_account_parent(name) is None]
        parents = sorted(parents)
        
        for p in parents:
            bal = get_account_balance(p)
            tree_rows.append({
                "Account Name": p,
                "Account Type": typ,
                "Level": "Main Ledger",
                "Parent Account": "",
                "Balance": f"₹{bal:,.2f}"
            })
            # Sub-ledgers for this parent
            children = [name for name in type_accs if get_account_parent(name) == p]
            children = sorted(children)
            for c in children:
                c_bal = get_account_balance(c)
                tree_rows.append({
                    "Account Name": f"　└── {c}",
                    "Account Type": typ,
                    "Level": "Sub-ledger",
                    "Parent Account": p,
                    "Balance": f"₹{c_bal:,.2f}"
                })
                
    if tree_rows:
        st.dataframe(pd.DataFrame(tree_rows), use_container_width=True, height=400, hide_index=True)
    else:
        st.info("No accounts found.")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 7 — ASSET PROJECTIONS
# ─────────────────────────────────────────────────────────────────────────────

with tab_fc:
    st.subheader("Asset & Networth Forecasting")
    
    # Make sure gold_qty exists in session_state
    if "gold_qty" not in st.session_state:
        st.session_state.gold_qty = 177.0
        
    # Calculate historical average monthly net balance (Net Receipts - Net Payout)
    unique_months = set()
    for jv in get_all_journal_entries():
        if hasattr(jv["date"], "strftime"):
            unique_months.add(jv["date"].strftime("%Y-%m"))
        else:
            unique_months.add(str(jv["date"])[:7])
    num_months = max(1, len(unique_months))
    historical_net_balance = (total_receipts - total_payout) / num_months

    # Layout forecasting parameters
    param_col1, param_col2, param_col3 = st.columns(3)
    with param_col1:
        gold_grams = st.number_input("Gold Quantity (grams)", min_value=0.0, value=float(st.session_state.gold_qty), step=1.0, format="%.2f", key="gold_qty_input")
        if gold_grams != st.session_state.gold_qty:
            st.session_state.gold_qty = gold_grams
            save_data()
            st.rerun()
    with param_col2:
        gold_depreciation = 0.23
        st.metric("Gold Depreciation", f"{gold_depreciation * 100:.0f}%", help="Fixed 23% depreciation from gross gold rate")
    with param_col3:
        # Adjustable monthly savings surplus (Net Balance growth)
        proj_net_balance = st.number_input("Projected Monthly Savings (Net Balance)", value=float(historical_net_balance), step=500.0, format="%.2f", help="Net savings surplus added to your liquid cash/bank assets each month (defaults to current month's Net Receipts minus Net Payout)")
        
    gold_rate = fetch_gold_price_inr()
    
    # Check if there is gold asset balance in ledger, otherwise fall back to physical weight pricing
    live_gold_bal = get_account_balance("Gold Asset")
    current_gold_value = live_gold_bal if live_gold_bal > 0 else (gold_grams * gold_rate * (1 - gold_depreciation))
    
    st.markdown(f"**Current Gold Assets Valuation:** ₹{current_gold_value:,.2f} (Rate: ₹{gold_rate:,.2f}/g)")
    
    # Identify dynamic asset and liability accounts
    asset_accounts = [name for name in st.session_state.accounts if get_account_type(name) == "Asset"]
    liab_accounts = [name for name in st.session_state.accounts if get_account_type(name) == "Liability"]
    
    # Starting values
    other_assets_start_sum = sum(get_account_balance(name) for name in asset_accounts if name != "Gold Asset")
    
    # We only include liability accounts in the projections table if they have a non-zero starting balance
    active_liab_accounts = sorted([name for name in liab_accounts if get_account_balance(name) != 0.0])
    
    # Precompute loan parameter dicts
    loan_params = {}
    for name in active_liab_accounts:
        starting_bal = get_account_balance(name)
        
        # 1. Gold Loans (contains "gold" and "loan" but not "interest" or "accrued")
        if "gold" in name.lower() and "loan" in name.lower() and "interest" not in name.lower() and "accrued" not in name.lower():
            rate = 9.5
            if "pnb" in name.lower() or "punjab" in name.lower():
                rate = 9.25
            elif "ksfe" in name.lower():
                rate = 9.85
            start_dt = get_account_start_date(name)
            loan_params[name] = {
                "type": "gold_loan",
                "starting_balance": starting_bal,
                "rate": rate,
                "start_date": start_dt
            }
            
        # 2. Amortizing Loans (contains "bajaj" or "emi")
        elif "bajaj" in name.lower() or "emi" in name.lower():
            monthly_pay = get_monthly_payment(name)
            loan_params[name] = {
                "type": "amortizing",
                "starting_balance": starting_bal,
                "monthly_payment": monthly_pay
            }
            
        # 3. Static Liabilities (Chitty, Accruals, Credit Cards, etc.)
        else:
            loan_params[name] = {
                "type": "static",
                "starting_balance": starting_bal
            }
            
    # Forecast chronologically 24 months forward
    start_date = datetime.now().replace(day=1)
    months_to_forecast = 24
    
    records = []
    
    for i in range(months_to_forecast):
        current_date = start_date + relativedelta(months=i)
        
        # Cumulative monthly net balance increase
        cumulative_savings = proj_net_balance * i
        total_assets_proj = other_assets_start_sum + current_gold_value + cumulative_savings
        
        # Project each liability account
        total_liabilities_proj = 0.0
        projected_liabs = {}
        
        for name, params in loan_params.items():
            starting_bal = params["starting_balance"]
            
            if params["type"] == "gold_loan":
                start_dt = params["start_date"]
                calc_current = datetime(current_date.year, current_date.month, 1)
                calc_start = datetime(start_dt.year, start_dt.month, 1)
                loan_months = max(0, (calc_current.year - calc_start.year) * 12 + (calc_current.month - calc_start.month))
                projected_val = starting_bal + (starting_bal * (params["rate"] / 100) * (loan_months / 12))
            
            elif params["type"] == "amortizing":
                if name.lower() == "bajaj loan":
                    # Special decomposition for combined Bajaj Loan based on the user's loan schedules.
                    # Since June payments are already posted, the remaining payments after June 2026 are:
                    # Loan 1: EMI 3152, 5 months remaining (expires Nov 2026)
                    # Loan 2: EMI 1000, 5 months remaining (expires Nov 2026)
                    # Loan 3: EMI 1688, 5 months remaining (expires Nov 2026)
                    # Loan 4: EMI 1750, 1 month remaining (expires July 2026)
                    # Loan 5: EMI 951, 2 months remaining (expires Aug 2026)
                    sub_loans = [
                        {"emi": 3152.0, "rem": 5},
                        {"emi": 1000.0, "rem": 5},
                        {"emi": 1688.0, "rem": 5},
                        {"emi": 1750.0, "rem": 1},
                        {"emi": 951.0, "rem": 2}
                    ]
                    total_paid = 0.0
                    for k in range(1, i + 1):
                        for sl in sub_loans:
                            if k <= sl["rem"]:
                                total_paid += sl["emi"]
                    projected_val = max(0.0, starting_bal - total_paid)
                else:
                    # General amortizing loan projection
                    projected_val = max(0.0, starting_bal - params["monthly_payment"] * i)
                
            else: # static
                projected_val = starting_bal
                
            projected_liabs[name] = projected_val
            total_liabilities_proj += projected_val
            
        networth_proj = total_assets_proj - total_liabilities_proj
        
        record = {
            "Date": current_date.strftime("%Y-%m"),
            "Assets (Proj)": total_assets_proj,
            "Total Liabilities (Proj)": total_liabilities_proj,
            "Networth": networth_proj
        }
        # Add dynamic liability columns
        for name in active_liab_accounts:
            record[name] = projected_liabs[name]
            
        records.append(record)
        
    df_forecast = pd.DataFrame(records)
    
    fig_nw = px.area(df_forecast, x="Date", y=["Total Liabilities (Proj)", "Assets (Proj)", "Networth"], 
                     title="24-Month Projections", template="none")
    fig_nw.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", font=dict(color="#0F172A"))
    st.plotly_chart(fig_nw, use_container_width=True)
    
    st.dataframe(df_forecast.style.format(formatter={col: "₹{:,.2f}" for col in df_forecast.columns if col != "Date"}), use_container_width=True, height=300, hide_index=True)

# ─────────────────────────────────────────────────────────────────────────────
# TAB 8 — ANALYTICS
# ─────────────────────────────────────────────────────────────────────────────
_DARK_BG   = "rgba(0,0,0,0)"
_PURPLE    = "#10B981"
_GREEN     = "#059669"
_RED       = "#EF4444"
_YELLOW    = "#F59E0B"
_PURPLES   = ["#D1FAE5","#A7F3D0","#6EE7B7","#34D399","#10B981","#059669","#047857","#065F46"]

def _chart_layout(fig, height=400, legend_h=False):
    fig.update_layout(
        paper_bgcolor=_DARK_BG,
        plot_bgcolor=_DARK_BG,
        font=dict(color="#0F172A", family="Outfit, Plus Jakarta Sans, sans-serif"),
        height=height,
        margin=dict(t=50, b=30, l=10, r=10),
        legend=dict(
            orientation="h" if legend_h else "v",
            bgcolor="rgba(0,0,0,0)",
            font=dict(color="#475569"),
            y=1.08 if legend_h else 1,
        ),
        xaxis=dict(gridcolor="rgba(0,0,0,0.05)", tickfont=dict(color="#475569")),
        yaxis=dict(gridcolor="rgba(0,0,0,0.05)", tickfont=dict(color="#475569")),
    )
    return fig

with tab_an:
    st.subheader("Analytics Dashboard")
    
    if not get_all_journal_entries():
        st.info("No transaction data loaded. Post JVs to see dynamic analytics visualizations.")
    else:
        # Build base analytics grid representation
        monthly_data = []
        for jv in get_all_journal_entries():
            jv_date = pd.to_datetime(jv["date"])
            month_str = jv_date.strftime("%Y-%m")
            for l in jv["lines"]:
                acc_name = l["account"]
                acc_type = get_account_type(acc_name)
                monthly_data.append({
                    "Date": jv_date,
                    "Month": month_str,
                    "MonthSort": jv_date.to_period("M"),
                    "Account": acc_name,
                    "Type": acc_type,
                    "Debit": l["debit"],
                    "Credit": l["credit"]
                })
        df_adf = pd.DataFrame(monthly_data)
        
        # Calculate revenues and expenses legs
        df_adf["Revenue"] = df_adf.apply(lambda r: r["Credit"] - r["Debit"] if r["Type"] == "Revenue" else 0.0, axis=1)
        df_adf["Expense"] = df_adf.apply(lambda r: r["Debit"] - r["Credit"] if r["Type"] == "Expense" else 0.0, axis=1)
        
        _total_income  = df_adf["Revenue"].sum()
        _total_expense = df_adf["Expense"].sum()
        _net_savings   = _total_income - _total_expense
        _savings_rate  = round((_net_savings / _total_income * 100), 1) if _total_income > 0 else 0.0
        _expense_ratio = round((_total_expense / _total_income * 100), 1) if _total_income > 0 else 0.0

        st.markdown("#### 📌 Key Performance Indicators")
        kpi1, kpi2, kpi3, kpi4 = st.columns(4)
        kpi1.metric("Total Period Revenue", f"₹{_total_income:,.0f}")
        kpi2.metric("Total Period Expenses", f"₹{_total_expense:,.0f}")
        kpi3.metric("Net Surplus", f"₹{_net_savings:,.0f}", delta=f"₹{_net_savings:,.0f}", delta_color="normal" if _net_savings >= 0 else "inverse")
        kpi4.metric("Surplus Savings Rate", f"{_savings_rate}%")

        st.markdown("---")

        # Monthly Income vs Expense Bar Chart
        st.markdown("#### 📅 Monthly Revenue vs Expense")
        df_monthly = df_adf.groupby("Month")[["Revenue", "Expense"]].sum().reset_index()
        df_monthly_melt = df_monthly.melt(id_vars="Month", value_vars=["Revenue", "Expense"], var_name="Type", value_name="Amount")
        
        fig_bar = px.bar(
            df_monthly_melt, x="Month", y="Amount", color="Type",
            barmode="group",
            color_discrete_map={"Revenue": _GREEN, "Expense": _RED},
            template="none",
            labels={"Amount": "Amount (₹)", "Month": "Month", "Type": ""},
        )
        fig_bar.update_traces(marker_line_width=0)
        _chart_layout(fig_bar, height=350, legend_h=True)
        st.plotly_chart(fig_bar, use_container_width=True)

        st.markdown("---")

        st.markdown("#### 🔍 Expense Account Breakdown & Liquid Assets")
        ch_left, ch_right = st.columns(2)

        with ch_left:
            # Expense by account
            df_exp_grouped = df_adf[df_adf["Type"] == "Expense"].groupby("Account")["Expense"].sum().reset_index()
            df_exp_grouped = df_exp_grouped[df_exp_grouped["Expense"] > 0].sort_values("Expense", ascending=False)
            
            if not df_exp_grouped.empty:
                fig_donut = px.pie(
                    df_exp_grouped, names="Account", values="Expense",
                    hole=0.58,
                    color_discrete_sequence=_PURPLES,
                    template="none"
                )
                fig_donut.update_traces(
                    textposition="inside",
                    textinfo="percent",
                    hovertemplate="<b>%{label}</b><br>₹%{value:,.0f}<br>%{percent}<extra></extra>"
                )
                fig_donut.add_annotation(
                    text=f"<b>₹{_total_expense:,.0f}</b><br><span style='font-size:11px'>Total Expense</span>",
                    x=0.5, y=0.5, showarrow=False,
                    font=dict(size=14, color="#0F172A"),
                    align="center"
                )
                fig_donut.update_layout(
                    paper_bgcolor=_DARK_BG,
                    font=dict(color="#0F172A"),
                    height=350,
                    margin=dict(t=30, b=10, l=10, r=10),
                    showlegend=True,
                    legend=dict(
                        bgcolor="rgba(0,0,0,0)",
                        font=dict(color="#475569", size=11),
                    )
                )
                st.plotly_chart(fig_donut, use_container_width=True)
            else:
                st.info("No expense legs recorded to chart.")

        with ch_right:
            # Running Liquid Asset Balance (Cash + Bank Account)
            df_liq = df_adf[df_adf["Account"].isin(["Cash", "Punjab National Bank", "Jio Payment Bank"])].copy()
            if not df_liq.empty:
                df_liq = df_liq.sort_values("Date")
                df_liq["NetChange"] = df_liq["Debit"] - df_liq["Credit"]
                df_liq["Running Balance"] = df_liq["NetChange"].cumsum()
                
                fig_line = px.line(
                    df_liq, x="Date", y="Running Balance",
                    color_discrete_sequence=[_PURPLE],
                    template="none",
                    labels={"Running Balance": "Liquid Balance (₹)", "Date": ""}
                )
                fig_line.update_traces(
                    fill="tozeroy",
                    fillcolor="rgba(16,185,129,0.15)",
                    line=dict(width=2.5),
                    hovertemplate="<b>%{x|%d %b %Y}</b><br>₹%{y:,.0f}<extra></extra>"
                )
                _chart_layout(fig_line, height=350)
                st.plotly_chart(fig_line, use_container_width=True)
            else:
                st.info("No Cash, Punjab National Bank, or Jio Payment Bank activity logged.")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 9 — TEMPLATE SYNC (IMPORT / EXPORT)
# ─────────────────────────────────────────────────────────────────────────────
with tab_io:
    st.subheader("Data Interchange & Import Sync")
    
    def generate_journal_template():
        wb = Workbook()
        ws = wb.active
        ws.title = "Journal"
        ws.append(["JV ID", "Date", "Narration", "Account", "Debit", "Credit"])
        
        # Validation lists
        ws_val = wb.create_sheet(title="Validation_Accounts")
        accs_list = list(st.session_state.accounts.keys())
        for idx, acc in enumerate(accs_list, start=1):
            ws_val.cell(row=idx, column=1, value=acc)
            
        dv = DataValidation(type="list", formula1=f"=Validation_Accounts!$A$1:$A${len(accs_list)}", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add("D2:D1000") # Column D is Account
        
        # Write existing records
        for jv in get_all_journal_entries():
            d_val = jv["date"].strftime("%Y-%m-%d") if hasattr(jv["date"], "strftime") else str(jv["date"])
            for l in jv["lines"]:
                ws.append([jv["jv_id"], d_val, jv["narration"], l["account"], l["debit"], l["credit"]])
                
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def generate_accounts_template():
        wb = Workbook()
        ws = wb.active
        ws.title = "Accounts"
        ws.append(["Account Name", "Account Type", "Parent Account"])
        for name, info in st.session_state.accounts.items():
            parent = get_account_parent(name)
            typ = get_account_type(name)
            ws.append([name, typ, parent or ""])
            
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    col_t1, col_t2 = st.columns(2)
    
    with col_t1:
        st.markdown("### 📥 Download Spreadsheet Templates")
        st.download_button(
            label="📥 Download Journal Ledger (.xlsx)",
            data=generate_journal_template(),
            file_name="journal_ledger_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        st.download_button(
            label="📥 Download Chart of Accounts (.xlsx)",
            data=generate_accounts_template(),
            file_name="chart_of_accounts.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

        st.markdown("---")
        st.markdown("### 🔄 Sync Trading P&L (FinPlus sqlite3 db)")
        st.markdown("""
        <div style="font-size:0.8rem;color:#64748B;font-weight:500;margin-bottom:0.8rem;">
            Pulls all-time Net Realized P&L from trading database and logs a balanced synced JV entry.
        </div>
        """, unsafe_allow_html=True)
        
        if st.button("🔄 Sync Net PNL Ledger Entry", use_container_width=True):
            import sqlite3
            db_path = r"C:\Users\AbhilashBabu\Finance\trading_journal.db"
            try:
                conn = sqlite3.connect(db_path)
                c = conn.cursor()
                
                c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='trades';")
                p1 = 0.0
                if c.fetchone():
                    c.execute("SELECT SUM(net_pnl) FROM trades")
                    p1 = c.fetchone()[0] or 0.0
                    
                c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='pair_trades';")
                p2 = 0.0
                if c.fetchone():
                    c.execute("SELECT SUM(net_pnl) FROM pair_trades")
                    p2 = c.fetchone()[0] or 0.0
                    
                conn.close()
                total_pnl = p1 + p2
                
                # Delete existing sync entries to avoid duplicates
                st.session_state.journal_entries = [
                    jv for jv in st.session_state.journal_entries 
                    if jv["narration"] != "Synced FinPlus PNL"
                ]
                
                if "Stock Market Asset" not in st.session_state.accounts:
                    st.session_state.accounts["Stock Market Asset"] = {"type": "Asset", "parent": None}
                    
                next_id = f"JV-{len(st.session_state.journal_entries) + 1:05d}"
                lines = []
                
                if total_pnl >= 0:
                    if "Stock Market Gains" not in st.session_state.accounts:
                        st.session_state.accounts["Stock Market Gains"] = {"type": "Revenue", "parent": None}
                    lines.append({"account": "Stock Market Asset", "debit": abs(total_pnl), "credit": 0.0})
                    lines.append({"account": "Stock Market Gains", "debit": 0.0, "credit": abs(total_pnl)})
                else:
                    if "Stock Market Loss" not in st.session_state.accounts:
                        st.session_state.accounts["Stock Market Loss"] = {"type": "Expense", "parent": None}
                    lines.append({"account": "Stock Market Loss", "debit": abs(total_pnl), "credit": 0.0})
                    lines.append({"account": "Stock Market Asset", "debit": 0.0, "credit": abs(total_pnl)})
                    
                st.session_state.journal_entries.append({
                    "jv_id": next_id,
                    "date": datetime.now().date(),
                    "narration": "Synced FinPlus PNL",
                    "lines": lines
                })
                save_data()
                st.session_state.success_msg = f"✅ Sync successful! Net Realized P&L logged: ₹{total_pnl:,.2f} inside {next_id}"
                st.rerun()
            except Exception as e:
                st.error(f"⚠️ Sync unsuccessful: {e}")

    with col_t2:
        st.markdown("### 📤 Upload Completed Spreadsheet")
        uploaded_file = st.file_uploader("Upload Excel Template (.xlsx)", type=["xlsx"])
        
        if uploaded_file is not None:
            st.session_state["_pending_ledger_upload"] = uploaded_file.read()
            st.session_state["_pending_ledger_name"] = uploaded_file.name
            
        if st.session_state.get("_pending_ledger_upload") is not None:
            st.info(f"📄 Staged for import: **{st.session_state['_pending_ledger_name']}**")
            
            imp_c, disc_c = st.columns([3, 1])
            with imp_c:
                if st.button("✅ Confirm Import & Sync", use_container_width=True):
                    try:
                        excel_data = pd.ExcelFile(io.BytesIO(st.session_state["_pending_ledger_upload"]))
                        imported = False
                        
                        # Parse Accounts
                        if "Accounts" in excel_data.sheet_names:
                            acc_df = excel_data.parse("Accounts")
                            if "Account Name" in acc_df.columns and "Account Type" in acc_df.columns:
                                has_parent = "Parent Account" in acc_df.columns
                                for _, row in acc_df.dropna(subset=["Account Name", "Account Type"]).iterrows():
                                    name = str(row["Account Name"]).strip()
                                    typ = str(row["Account Type"]).strip()
                                    parent = str(row["Parent Account"]).strip() if (has_parent and pd.notnull(row["Parent Account"])) else ""
                                    if parent == "nan" or not parent:
                                        parent = None
                                    if typ in ["Asset", "Liability", "Equity", "Revenue", "Expense"]:
                                        st.session_state.accounts[name] = {"type": typ, "parent": parent}
                                imported = True
                                
                        # Parse Journal Transactions
                        if "Journal" in excel_data.sheet_names:
                            journal_df = excel_data.parse("Journal")
                            required_cols = ["JV ID", "Date", "Narration", "Account", "Debit", "Credit"]
                            
                            if all(col in journal_df.columns for col in required_cols):
                                journal_df["Debit"] = pd.to_numeric(journal_df["Debit"], errors="coerce").fillna(0.0)
                                journal_df["Credit"] = pd.to_numeric(journal_df["Credit"], errors="coerce").fillna(0.0)
                                journal_df["Date"] = pd.to_datetime(journal_df["Date"], errors="coerce").dt.date
                                journal_df["Account"] = journal_df["Account"].astype(str).str.strip()
                                journal_df["Narration"] = journal_df["Narration"].fillna("").astype(str)
                                journal_df["JV ID"] = journal_df["JV ID"].astype(str).str.strip()
                                
                                imported_jvs = []
                                unbalanced_jvs = []
                                grouped = journal_df.groupby("JV ID")
                                
                                for jv_id, group in grouped:
                                    dr_sum = group["Debit"].sum()
                                    cr_sum = group["Credit"].sum()
                                    if abs(dr_sum - cr_sum) > 0.01:
                                        unbalanced_jvs.append(jv_id)
                                        continue
                                        
                                    jv_date_val = group["Date"].iloc[0]
                                    if pd.isnull(jv_date_val):
                                        jv_date_val = datetime.now().date()
                                        
                                    lines = []
                                    for _, row in group.iterrows():
                                        if row["Debit"] > 0 or row["Credit"] > 0:
                                            acc = row["Account"]
                                            if acc not in st.session_state.accounts:
                                                st.session_state.accounts[acc] = {"type": "Expense", "parent": None}
                                            lines.append({
                                                "account": acc,
                                                "debit": float(row["Debit"]),
                                                "credit": float(row["Credit"])
                                            })
                                            
                                    if lines:
                                        imported_jvs.append({
                                            "jv_id": jv_id,
                                            "date": jv_date_val,
                                            "narration": group["Narration"].iloc[0],
                                            "lines": lines
                                        })
                                        
                                if unbalanced_jvs:
                                    st.warning(f"⚠️ Unbalanced JVs skipped: {', '.join(unbalanced_jvs)}")
                                if imported_jvs:
                                    st.session_state.journal_entries = imported_jvs
                                    imported = True
                                    
                        st.session_state["_pending_ledger_upload"] = None
                        st.session_state["_pending_ledger_name"] = None
                        
                        if imported:
                            st.session_state.success_msg = "✅ Excel data synced successfully!"
                            save_data()
                            st.rerun()
                        else:
                            st.error("No valid matching templates found (sheets named 'Journal' or 'Accounts').")
                    except Exception as e:
                        st.error(f"Failed to read file: {e}")
                        
            with disc_c:
                if st.button("Discard", use_container_width=True):
                    st.session_state["_pending_ledger_upload"] = None
                    st.session_state["_pending_ledger_name"] = None
                    st.rerun()


# ─────────────────────────────────────────────────────────────────────────────
# TAB 10 — IMPORTANT EVENTS & DOCUMENT VAULT
# ─────────────────────────────────────────────────────────────────────────────
with tab_ev:
    DOCS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "event_documents")
    os.makedirs(DOCS_DIR, exist_ok=True)
    
    st.markdown('<div class="glass-card preview-card">', unsafe_allow_html=True)
    st.markdown('<h3 style="margin-top:0; color: var(--text-primary);">🚗 Important Events & Document Vault</h3>', unsafe_allow_html=True)
    st.markdown('<p style="color: var(--text-secondary); margin-top:-10px;">Mark important milestones (e.g., buying a car, property purchase, major expenses) and upload bills, receipts, invoices, or photos. Everything stays saved securely in your local vault.</p>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    events = st.session_state.get("important_events", [])
    total_events = len(events)
    total_docs = sum(len(e.get("documents", [])) for e in events)
    
    kpi_col1, kpi_col2, kpi_col3 = st.columns(3)
    with kpi_col1:
        st.metric("Total Tracked Events", total_events)
    with kpi_col2:
        st.metric("Stored Invoices & Documents", total_docs)
    with kpi_col3:
        dir_size_mb = 0.0
        if os.path.exists(DOCS_DIR):
            for f in os.listdir(DOCS_DIR):
                fp = os.path.join(DOCS_DIR, f)
                if os.path.isfile(fp):
                    dir_size_mb += os.path.getsize(fp) / (1024 * 1024)
        st.metric("Vault Storage Used", f"{dir_size_mb:.2f} MB")

    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    st.markdown("<h4>📝 Log Event or Add Document</h4>", unsafe_allow_html=True)
    
    event_names = sorted(list(set(e["name"] for e in events))) if events else []
    event_options = ["-- Create New Event --"] + event_names
    
    selected_event_opt = st.selectbox("Link to Event", options=event_options, index=0)
    
    ev_col1, ev_col2 = st.columns(2)
    
    with ev_col1:
        if selected_event_opt == "-- Create New Event --":
            new_event_name = st.text_input("New Event Name", value="", placeholder="e.g. Bought Nexon EV Car, House Renovation")
            event_date = st.date_input("Event Date", value=datetime.now().date())
        else:
            existing_event = next(e for e in events if e["name"] == selected_event_opt)
            new_event_name = selected_event_opt
            try:
                event_date_val = datetime.strptime(existing_event["event_date"], "%Y-%m-%d").date()
            except Exception:
                event_date_val = datetime.now().date()
            event_date = st.date_input("Event Date", value=event_date_val, disabled=True)
            
        doc_type = st.selectbox("Document Type", options=["Invoice / Bill", "Photo / Image", "Agreement / Certificate", "Insurance Policy", "Other"])
        
    with ev_col2:
        uploaded_file = st.file_uploader("Upload File (PDF, PNG, JPG, Excel, etc.)", type=["pdf", "png", "jpg", "jpeg", "xlsx", "docx", "txt", "zip"])
        event_notes = st.text_area("Description / Notes", placeholder="Add details or narrations for this event/document...", height=100)

    if st.button("🚀 Upload & Save to Vault", use_container_width=True):
        if not new_event_name.strip():
            st.error("Please enter a valid Event Name.")
        else:
            doc_info = None
            if uploaded_file is not None:
                import uuid
                file_ext = os.path.splitext(uploaded_file.name)[1]
                disk_filename = f"{uuid.uuid4().hex}{file_ext}"
                disk_filepath = os.path.join(DOCS_DIR, disk_filename)
                
                try:
                    with open(disk_filepath, "wb") as f:
                        f.write(uploaded_file.getbuffer())
                    
                    doc_info = {
                        "id": f"DOC-{uuid.uuid4().hex[:8].upper()}",
                        "filename": disk_filename,
                        "original_filename": uploaded_file.name,
                        "doc_type": doc_type,
                        "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "size_kb": round(len(uploaded_file.getbuffer()) / 1024, 2)
                    }
                except Exception as e:
                    st.error(f"Failed to save file: {e}")
                    st.stop()
                    
            existing_event_idx = None
            for idx, e in enumerate(events):
                if e["name"] == new_event_name:
                    existing_event_idx = idx
                    break
                    
            if existing_event_idx is not None:
                if doc_info:
                    events[existing_event_idx]["documents"].append(doc_info)
                if event_notes.strip():
                    old_notes = events[existing_event_idx].get("notes", "")
                    events[existing_event_idx]["notes"] = f"{old_notes}\n{event_notes}".strip()
                st.success(f"Successfully added document to event: **{new_event_name}**!")
            else:
                import uuid
                new_event = {
                    "id": f"EV-{uuid.uuid4().hex[:8].upper()}",
                    "name": new_event_name,
                    "event_date": event_date.strftime("%Y-%m-%d"),
                    "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "documents": [doc_info] if doc_info else [],
                    "notes": event_notes.strip()
                }
                events.append(new_event)
                st.success(f"Successfully logged new event: **{new_event_name}**!")
                
            st.session_state.important_events = events
            save_data()
            st.rerun()
            
    st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    st.markdown("<h3>📂 Document Vault & Event Ledger</h3>", unsafe_allow_html=True)
    
    if not events:
        st.info("No events or documents logged yet. Use the form above to record your first milestone!")
    else:
        search_q = st.text_input("Search Events or Documents", placeholder="e.g. Nexon, Invoice, 2026").strip().lower()
        
        sorted_events = sorted(events, key=lambda x: x.get("event_date", ""), reverse=True)
        
        for e in sorted_events:
            match = False
            if not search_q:
                match = True
            else:
                if search_q in e["name"].lower() or search_q in e.get("notes", "").lower() or search_q in e.get("event_date", ""):
                    match = True
                else:
                    for doc in e.get("documents", []):
                        if search_q in doc["original_filename"].lower() or search_q in doc["doc_type"].lower():
                            match = True
                            break
            
            if not match:
                continue
                
            st.markdown(
                f"""
                <div style="border: 1px solid #E2E8F0; border-left: 5px solid #059669; border-radius: 8px; padding: 15px; margin-bottom: 15px; background: #FFFFFF; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.02);">
                    <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px;">
                        <span style="font-weight:800; font-size:1.15rem; color: #0F172A;">{e['name']}</span>
                        <span style="background-color: #ECFDF5; color: #059669; padding: 3px 8px; border-radius: 6px; font-weight: 700; font-size: 0.85rem;">
                            📅 {e['event_date']}
                        </span>
                    </div>
                    <div style="font-size: 0.9rem; color: #475569; margin-bottom: 10px; white-space: pre-line;">
                        {e.get('notes', 'No notes added.')}
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )
            
            docs = e.get("documents", [])
            
            if not docs:
                st.markdown("<p style='font-size:0.85rem; color: #94A3B8; margin-left: 20px;'>No attachments or bills uploaded.</p>", unsafe_allow_html=True)
            else:
                st.markdown("<h6 style='margin: 0 0 8px 20px; font-size:0.9rem; color: #334155; font-weight: 700;'>Attachments & Receipts:</h6>", unsafe_allow_html=True)
                
                for doc in docs:
                    doc_id = doc["id"]
                    disk_file = doc["filename"]
                    orig_file = doc["original_filename"]
                    doc_t = doc["doc_type"]
                    u_date = doc["uploaded_at"]
                    sz = doc["size_kb"]
                    
                    doc_filepath = os.path.join(DOCS_DIR, disk_file)
                    
                    ext = os.path.splitext(orig_file)[1].lower()
                    if ext in [".png", ".jpg", ".jpeg"]:
                        file_icon = "🖼️"
                    elif ext in [".pdf"]:
                        file_icon = "📄"
                    elif ext in [".xlsx", ".xls"]:
                        file_icon = "📊"
                    else:
                        file_icon = "📝"
                        
                    doc_col1, doc_col2, doc_col3 = st.columns([6, 2, 2])
                    
                    with doc_col1:
                        st.markdown(
                            f"""
                            <div style="margin-left: 20px; font-size: 0.85rem; display: flex; align-items: center; gap: 8px;">
                                <span style="font-size: 1.1rem;">{file_icon}</span>
                                <div>
                                    <strong>{orig_file}</strong> <span style="color: #64748B;">({doc_t} // {sz} KB)</span><br>
                                    <span style="font-size:0.75rem; color:#94A3B8;">Uploaded on: {u_date}</span>
                                </div>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )
                        
                    file_data = b""
                    file_exists = os.path.exists(doc_filepath)
                    if file_exists:
                        try:
                            with open(doc_filepath, "rb") as f_read:
                                file_data = f_read.read()
                        except Exception:
                            pass
                            
                    with doc_col2:
                        if file_exists and file_data:
                            st.download_button(
                                label="📥 Download",
                                data=file_data,
                                file_name=orig_file,
                                key=f"dl_{doc_id}",
                                use_container_width=True
                            )
                        else:
                            st.button("⚠️ Missing file", disabled=True, key=f"missing_{doc_id}", use_container_width=True)
                            
                    with doc_col3:
                        if st.button("🗑️ Delete File", key=f"del_doc_{doc_id}", use_container_width=True):
                            if file_exists:
                                try:
                                    os.remove(doc_filepath)
                                except Exception:
                                    pass
                            e["documents"] = [d for d in e["documents"] if d["id"] != doc_id]
                            st.success(f"Deleted {orig_file}")
                            save_data()
                            st.rerun()
                            
                    st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)
            
            del_ev_col1, del_ev_col2 = st.columns([8, 2])
            with del_ev_col2:
                if st.button("🗑️ Delete Event", key=f"del_ev_{e['id']}", use_container_width=True):
                    for doc in e.get("documents", []):
                        dfp = os.path.join(DOCS_DIR, doc["filename"])
                        if os.path.exists(dfp):
                            try:
                                os.remove(dfp)
                            except Exception:
                                pass
                    st.session_state.important_events = [ev for ev in events if ev["id"] != e["id"]]
                    st.warning(f"Deleted event: {e['name']}")
                    save_data()
                    st.rerun()
            st.markdown("<hr style='border-color: #E2E8F0; margin: 15px 0 20px 0;'>", unsafe_allow_html=True)
            
    st.markdown('</div>', unsafe_allow_html=True)

